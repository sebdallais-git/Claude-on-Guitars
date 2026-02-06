#!/usr/bin/env python3
"""
searcher — crawls vintage guitar sites every 5 minutes.

Supports:
- retrofret.com (USA, USD)
- woodstore.fr (France, EUR)

Filters by condition >= excellent-.  Skips "on hold" items.
Writes results to an Excel sheet, appending new hits sorted by ProductID.
Caches conditions so each product page is fetched only once.
"""

import re
import os
import sys
import time
import json
from datetime import datetime, date
from concurrent.futures import ThreadPoolExecutor

try:
    import requests
    from bs4 import BeautifulSoup
except ImportError:
    print("Missing deps.  pip install requests beautifulsoup4 openpyxl")
    sys.exit(1)

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Import woodstore scraper
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "scrapers"))
try:
    import woodstore
    WOODSTORE_ENABLED = True
except ImportError:
    WOODSTORE_ENABLED = False
    print("  [!] woodstore.fr scraper not available")

# ── config ───────────────────────────────────────────────────────
BASE_URL   = "https://www.retrofret.com"
CATEGORIES = ["/acoustic/", "/electric/", "/bass/"]

_SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.dirname(_SCRIPT_DIR)
_DATA         = os.path.join(_PROJECT_ROOT, "data")
_OUTPUTS      = os.path.join(_PROJECT_ROOT, "outputs")

OUTPUT_FILE          = os.path.join(_OUTPUTS, "listings.xlsx")
CACHE_FILE           = os.path.join(_DATA, ".condition_cache.json")
REVERB_NO_DATA_FILE  = os.path.join(_DATA, ".reverb-no-data.json")
SOLD_CANDIDATES_FILE = os.path.join(_DATA, ".sold_candidates.json")
INTERVAL             = 300  # seconds
SOLD_THRESHOLD       = 600  # seconds — 2 scrape cycles before confirming sold

HTTP_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    )
}

# ── condition ────────────────────────────────────────────────────
CONDITION_SCALE = [
    "poor", "fair",
    "good-", "good", "good+",
    "very good-", "very good", "very good+",
    "excellent-", "excellent", "excellent+",
    "near mint", "mint",
]
MIN_CONDITION_RANK = CONDITION_SCALE.index("excellent-")


def condition_rank(cond):
    if cond is None:
        return -1
    try:
        return CONDITION_SCALE.index(cond.lower())
    except ValueError:
        return -1


def parse_condition(page_text):
    m = re.search(
        r"(?:overall\s+)?"
        r"(mint|near\s*mint|excellent|very\s+good|good|fair|poor)"
        r"\s*([+-])?"
        r"\s*(?:condition)",
        page_text, re.IGNORECASE,
    )
    if not m:
        return None
    base = re.sub(r"\s+", " ", m.group(1).strip().lower())
    return f"{base}{m.group(2) or ''}"


# ── title parsing ────────────────────────────────────────────────
# Brands ordered longest-first to prevent partial matches
KNOWN_BRANDS = [
    "C. F. Martin", "C.F. Martin",
    "Kolya Panhuyzen", "Recording King", "Vicente Tatay",
    "Felipe Conde", "Jerry Jones", "Vivi-Tone",
    "D'Angelico", "D\u2019Angelico",
    "Gibson", "Fender", "Taylor", "Guild",
    "Epiphone", "Gretsch", "Rickenbacker",
    "Maccaferri", "Wandre", "Mosrite",
    "Ampeg", "Audiovox", "Hohner", "Martin",
]

# Marks the boundary between model name and type/extra text
_SPLIT_RE = re.compile(
    r"\b(?:"
    r"Flat Top|Arch Top|Solid Body|Hollow Body|Semi[- ]Hollow Body|"
    r"12[- ]?String|Tenor|Classical|Resophonic|Lap Steel|Steel Guitar|"
    r"Harp Guitar|Jumbo|Piggyback|Tube|"
    r"Acoustic|Electric|Bass Guitar|"
    r"Owned|Played|[Mm]ade by|[Ll]eft[- ]?[Hh]anded|"
    r"w/"
    r")",
    re.IGNORECASE,
)


def extract_brand_model(title):
    """Return (brand, model) parsed from the listing title."""
    clean = re.sub(r"[,(]\s*(?:c\.\s*)?\d{3,4}s?\s*\)?\s*$", "", title).strip()
    for brand in KNOWN_BRANDS:
        if clean.lower().startswith(brand.lower()):
            rest  = clean[len(brand):].strip()
            m     = _SPLIT_RE.search(rest)
            model = (rest[:m.start()] if m else rest).strip(" ,")
            return brand, model
    # fallback: first word is brand
    parts = clean.split(None, 1)
    brand = parts[0] if parts else ""
    rest  = parts[1] if len(parts) > 1 else ""
    m     = _SPLIT_RE.search(rest)
    model = (rest[:m.start()] if m else rest).strip(" ,")
    return brand, model


def extract_year(title):
    m = re.search(r"((?:c\.\s*)?\d{4}s?)", title)
    return m.group(1) if m else ""


def extract_type(title):
    t = title.lower()
    if "bass" in t:
        return "Bass"
    if "acoustic" in t:
        return "Acoustic"
    if "electric" in t:
        return "Electric"
    return "Unknown"


def extract_price(price_str):
    try:
        return float(price_str.replace("$", "").replace(",", ""))
    except (ValueError, AttributeError):
        return None


# ── I/O ──────────────────────────────────────────────────────────
def load_seen_ids():
    """Read ProductIDs already present in the Excel output."""
    seen = set()
    if os.path.exists(OUTPUT_FILE):
        wb = load_workbook(OUTPUT_FILE, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_col=11, values_only=True):
            url = str(row[10]) if row[10] else ""  # URL is now column 11 (index 10)
            # Handle both retrofret ProductID and woodstore slugs
            m = re.search(r"ProductID=(\d+)", url)
            if m:
                seen.add(m.group(1))
            else:
                # For woodstore and other sites, use the full URL as ID
                m = re.search(r"/products?/([^/?]+)", url)
                if m:
                    seen.add(m.group(1))
        wb.close()
    return seen


def load_cache():
    if os.path.exists(CACHE_FILE):
        with open(CACHE_FILE) as f:
            return json.load(f)
    return {}


def save_cache(cache):
    with open(CACHE_FILE, "w") as f:
        json.dump(cache, f)


def load_reverb_no_data():
    if os.path.exists(REVERB_NO_DATA_FILE):
        with open(REVERB_NO_DATA_FILE) as f:
            return set(json.load(f))
    return set()


def save_reverb_no_data(ids):
    with open(REVERB_NO_DATA_FILE, "w") as f:
        json.dump(sorted(ids), f)


def load_sold_candidates():
    if os.path.exists(SOLD_CANDIDATES_FILE):
        with open(SOLD_CANDIDATES_FILE) as f:
            return json.load(f)   # {pid: first_missed_timestamp}
    return {}


def save_sold_candidates(candidates):
    with open(SOLD_CANDIDATES_FILE, "w") as f:
        json.dump(candidates, f)


def load_sold_ids():
    """ProductIDs already marked as sold in the Excel."""
    sold = set()
    if os.path.exists(OUTPUT_FILE):
        wb = load_workbook(OUTPUT_FILE, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_col=13, values_only=True):
            if row[12] is not None:                          # col 13 = Sold Date
                url = str(row[10]) if row[10] else ""  # URL is now column 11 (index 10)
                m   = re.search(r"ProductID=(\d+)", url)
                if m:
                    sold.add(m.group(1))
        wb.close()
    return sold


# ── scraping ─────────────────────────────────────────────────────
def fetch_soup(url):
    try:
        resp = requests.get(url, headers=HTTP_HEADERS, timeout=15)
        resp.raise_for_status()
        return BeautifulSoup(resp.text, "html.parser")
    except requests.RequestException as e:
        print(f"  [!] {url} → {e}")
        return None


def max_page(soup):
    hi = 1
    for a in soup.find_all("a", href=re.compile(r"Page=\d+")):
        m = re.search(r"Page=(\d+)", a["href"])
        if m:
            hi = max(hi, int(m.group(1)))
    return hi


def clean_title(raw):
    year_m = re.search(r"(?:\(\s*(?:c\.\s*)?\d{3,4}s?\s*\)|,\s*c\.\s*\d{3,4})", raw)
    if year_m:
        return re.sub(r"\s+", " ", raw[:year_m.end()].strip())
    return re.sub(r"\s+", " ", re.split(r"(?:Just Arrived!|\$|Item\s*#)", raw)[0].strip())


def parse_listings(soup):
    guitars, seen = [], set()
    for tag in soup.find_all("a", href=re.compile(r"product\.asp\?ProductID=\d+")):
        pid = re.search(r"ProductID=(\d+)", tag["href"]).group(1)
        if pid in seen:
            continue
        title = clean_title(tag.get_text(strip=True))
        if not title:
            continue
        seen.add(pid)
        container      = tag.find_parent("li") or tag.find_parent("td") or tag
        container_text = container.get_text()
        on_hold        = bool(re.search(r"on hold", container_text, re.IGNORECASE))
        price_m        = re.search(r"\$[\d,]+(?:\.\d{2})?", container_text)
        guitars.append({
            "id":      pid,
            "title":   title,
            "price":   price_m.group(0) if price_m else "N/A",
            "url":     f"{BASE_URL}/{tag['href'].lstrip('/')}",
            "on_hold": on_hold,
        })
    return guitars


def scrape_all():
    """Scrape all configured sites and combine results."""
    all_g = {}

    # RetroFret
    for cat in CATEGORIES:
        soup = fetch_soup(f"{BASE_URL}{cat}")
        if not soup:
            continue
        pages = max_page(soup)
        for g in parse_listings(soup):
            g["source"] = "retrofret.com"
            all_g.setdefault(("retrofret", g["id"]), g)
        for p in range(2, pages + 1):
            time.sleep(0.3)
            soup = fetch_soup(f"{BASE_URL}{cat}Default.asp?Page={p}")
            if soup:
                for g in parse_listings(soup):
                    g["source"] = "retrofret.com"
                    all_g.setdefault(("retrofret", g["id"]), g)

    # Woodstore.fr
    if WOODSTORE_ENABLED:
        try:
            print("  also scraping woodstore.fr …")
            woodstore_guitars = woodstore.scrape_all()
            for g in woodstore_guitars:
                g["source"] = "woodstore.fr"
                # Woodstore guitars don't need condition fetching - assume good condition
                g.setdefault("condition", "excellent")
                all_g.setdefault(("woodstore", g["id"]), g)
            print(f"  + {len(woodstore_guitars)} from woodstore.fr")
        except Exception as e:
            print(f"  [!] woodstore error: {e}")

    return list(all_g.values())


def fetch_condition(url):
    try:
        resp = requests.get(url, headers=HTTP_HEADERS, timeout=15)
        resp.raise_for_status()
        return parse_condition(resp.text)
    except requests.RequestException:
        return None


# ── reverb price lookup ───────────────────────────────────────────
def _brands_match(query_brand, reverb_make):
    """Lenient brand match — 'C. F. Martin' matches 'Martin', etc."""
    a = re.sub(r"[^a-z]", "", query_brand.lower())
    b = re.sub(r"[^a-z]", "", reverb_make.lower())
    return bool(a and b and (a in b or b in a))


def reverb_price(brand, model, year):
    """Return (low, high) from Reverb price guide.  Tries multiple queries."""
    # clean model: strip leading punctuation (e.g. "/Gretsch ..." → "Gretsch ...")
    model = re.sub(r"^[^a-zA-Z0-9]+", "", model).strip()

    # model variants: full, then first-2-words if model is long
    models = [model]
    if len(model.split()) > 2:
        models.append(" ".join(model.split()[:2]))

    # brand variants: full, then last word if multi-word ("C. F. Martin" → "Martin")
    brands = [brand]
    short = brand.split()[-1] if brand else ""
    if short.lower() != brand.lower():
        brands.append(short)

    # build queries most-specific-first: with year, then without
    queries = []
    for y in ([year, ""] if year else [""]):
        for b in brands:
            for m in models:
                q = f"{b} {m} {y}".strip()
                if q and q not in queries:
                    queries.append(q)

    for query in queries:
        try:
            resp = requests.get(
                "https://api.reverb.com/api/priceguide",
                params={"query": query, "per_page": 10},
                headers={**HTTP_HEADERS,
                         "Accept": "application/hal+json",
                         "Accept-Version": "3.0"},
                timeout=10,
            )
            resp.raise_for_status()
            lows, highs = [], []
            for pg in resp.json().get("price_guides", []):
                ev = pg.get("estimated_value")
                if not ev or not _brands_match(brand, pg.get("make", "")):
                    continue
                lows.append(float(ev["price_low"]["amount"]))
                highs.append(float(ev["price_high"]["amount"]))
            if lows:
                return sum(lows) / len(lows), sum(highs) / len(highs)
        except Exception:
            continue
    return None, None


# ── Excel output ─────────────────────────────────────────────────
XLSX_HEADERS = [
    "Date Arrived", "Source", "Brand / Make", "Model",
    "Acoustic / Electric", "Year", "Price", "Reverb Low $", "Reverb High $",
    "Condition", "URL", "On Hold", "Sold Date",
]
XLSX_WIDTHS  = [14, 18, 22, 26, 18, 10, 14, 14, 14, 14, 65, 12, 14]
HEADER_FILL  = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
URL_FONT     = Font(color="0563C1", underline="single", size=10)


def _init_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Guitars"
    for col, (header, width) in enumerate(zip(XLSX_HEADERS, XLSX_WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 20
    return wb


def append_new(guitars, seen_ids):
    """Write new guitars to Excel; return set of new IDs."""
    new = sorted(
        (g for g in guitars if g["id"] not in seen_ids),
        key=lambda g: int(g["id"]),
    )
    if not new:
        print("  No new entries.\n")
        return set()

    # parse titles once, then look up Reverb prices in parallel
    parsed = [(extract_brand_model(g["title"]), extract_year(g["title"])) for g in new]
    print(f"  looking up Reverb prices for {len(new)} items …")
    with ThreadPoolExecutor(max_workers=3) as pool:
        futures    = [pool.submit(reverb_price, bm[0], bm[1], y) for bm, y in parsed]
        rev_prices = [f.result() for f in futures]

    if os.path.exists(OUTPUT_FILE):
        wb        = load_workbook(OUTPUT_FILE)
        ws        = wb.active
        ensure_sold_header(ws)
        start_row = ws.max_row + 1
    else:
        wb        = _init_workbook()
        ws        = wb.active
        start_row = 2

    today = date.today()
    for i, (g, ((brand, model), year), (rev_lo, rev_hi)) in enumerate(
        zip(new, parsed, rev_prices)
    ):
        row   = start_row + i
        gtype = extract_type(g["title"])
        price = extract_price(g["price"])
        cond  = g.get("condition") or "unknown"

        source = g.get("source", "retrofret.com")

        ws.cell(row=row, column=1, value=today)
        ws.cell(row=row, column=1).number_format = "YYYY-MM-DD"
        ws.cell(row=row, column=2, value=source)
        ws.cell(row=row, column=3, value=brand)
        ws.cell(row=row, column=4, value=model)
        ws.cell(row=row, column=5, value=gtype)
        ws.cell(row=row, column=6, value=year)
        if price is not None:
            ws.cell(row=row, column=7, value=price)
            # Use appropriate currency format based on source
            if "woodstore" in source:
                ws.cell(row=row, column=7).number_format = "€#,##0.00"
            else:
                ws.cell(row=row, column=7).number_format = "$#,##0.00"
        if rev_lo is not None:
            ws.cell(row=row, column=8, value=rev_lo)
            ws.cell(row=row, column=8).number_format = "$#,##0.00"
        if rev_hi is not None:
            ws.cell(row=row, column=9, value=rev_hi)
            ws.cell(row=row, column=9).number_format = "$#,##0.00"
        ws.cell(row=row, column=10, value=cond)
        url_cell          = ws.cell(row=row, column=11, value=g["url"])
        url_cell.font     = URL_FONT
        url_cell.hyperlink = g["url"]

        rev_str = f"  reverb ${rev_lo:,.0f}–${rev_hi:,.0f}" if rev_lo else ""
        print(f"  + {brand:22} {model:24} {year:>8}  {g['price']:>10}  [{cond}]{rev_str}")

    wb.save(OUTPUT_FILE)
    print()
    return {g["id"] for g in new}


def mark_on_hold(on_hold_ids):
    """Set 'On Hold' date for rows whose guitars are now on hold."""
    if not os.path.exists(OUTPUT_FILE):
        return
    wb      = load_workbook(OUTPUT_FILE)
    ws      = wb.active
    changed = False
    today   = date.today().isoformat()
    for row in range(2, ws.max_row + 1):
        url = str(ws.cell(row=row, column=11).value or "")  # URL is now column 11
        m   = re.search(r"ProductID=(\d+)", url)
        if m and m.group(1) in on_hold_ids and ws.cell(row=row, column=12).value is None:
            ws.cell(row=row, column=12, value=today)  # On Hold is now column 12
            changed = True
            print(f"  [on hold] {ws.cell(row=row, column=3).value}  "  # Brand is now column 3
                  f"{ws.cell(row=row, column=4).value}")  # Model is now column 4
    if changed:
        wb.save(OUTPUT_FILE)
        print()


def backfill_reverb(no_data_ids):
    """Retry Reverb lookups for any rows still missing prices.  Returns updated no_data set."""
    if not os.path.exists(OUTPUT_FILE):
        return no_data_ids
    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active
    to_fill = []
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=8).value is not None:  # Reverb Low is now column 8
            continue
        url = str(ws.cell(row=row, column=11).value or "")  # URL is now column 11
        m   = re.search(r"ProductID=(\d+)", url)
        pid = m.group(1) if m else str(row)
        if pid in no_data_ids:
            continue
        to_fill.append((row, pid,
                        str(ws.cell(row=row, column=3).value or ""),  # Brand is now column 3
                        str(ws.cell(row=row, column=4).value or ""),  # Model is now column 4
                        str(ws.cell(row=row, column=6).value or "")))  # Year is now column 6
    if not to_fill:
        return no_data_ids
    print(f"  backfilling Reverb for {len(to_fill)} entries …")
    with ThreadPoolExecutor(max_workers=3) as pool:
        futures = [pool.submit(reverb_price, b, m, y) for _, _, b, m, y in to_fill]
        results = [f.result() for f in futures]
    changed = False
    for (row, pid, brand, model, year), (lo, hi) in zip(to_fill, results):
        if lo is not None:
            ws.cell(row=row, column=8, value=lo)  # Reverb Low is now column 8
            ws.cell(row=row, column=8).number_format = "$#,##0.00"
            ws.cell(row=row, column=9, value=hi)  # Reverb High is now column 9
            ws.cell(row=row, column=9).number_format = "$#,##0.00"
            changed = True
            print(f"  [reverb] {brand} {model} → ${lo:,.0f} – ${hi:,.0f}")
        else:
            no_data_ids.add(pid)
    if changed:
        wb.save(OUTPUT_FILE)
    save_reverb_no_data(no_data_ids)
    return no_data_ids


# ── sold detection ───────────────────────────────────────────────
def check_sold(seen_ids, current_ids, sold_ids, candidates):
    """
    Grace-period sold detection.  A guitar is confirmed sold only after
    it has been absent from the site for >= SOLD_THRESHOLD seconds.
    Returns (updated_candidates, list_of_newly_confirmed_sold_pids).
    """
    now        = time.time()
    newly_sold = []
    for pid in seen_ids - sold_ids:
        if pid in current_ids:
            candidates.pop(pid, None)            # back on site — clear
        elif pid in candidates:
            if now - candidates[pid] >= SOLD_THRESHOLD:
                newly_sold.append(pid)
                del candidates[pid]
        else:
            candidates[pid] = now                # first miss — start grace
    return candidates, newly_sold


def ensure_sold_header(ws):
    """Add 'Sold Date' header to col 13 if missing (pre-existing workbooks)."""
    if ws.cell(row=1, column=13).value is None:
        cell           = ws.cell(row=1, column=13, value="Sold Date")
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(13)].width = 14


def mark_sold_batch(pids):
    """Write today's date into 'Sold Date' (col 13) for the given ProductIDs."""
    if not os.path.exists(OUTPUT_FILE) or not pids:
        return
    wb      = load_workbook(OUTPUT_FILE)
    ws      = wb.active
    ensure_sold_header(ws)
    today   = date.today().isoformat()
    pid_set = set(pids)
    for row in range(2, ws.max_row + 1):
        url = str(ws.cell(row=row, column=11).value or "")  # URL is now column 11
        m   = re.search(r"ProductID=(\d+)", url)
        if m and m.group(1) in pid_set and ws.cell(row=row, column=13).value is None:  # Sold Date is now column 13
            ws.cell(row=row, column=13, value=today)
            print(f"  [sold] {ws.cell(row=row, column=3).value}  "  # Brand is now column 3
                  f"{ws.cell(row=row, column=4).value}")  # Model is now column 4
    wb.save(OUTPUT_FILE)
    print()


# ── display ──────────────────────────────────────────────────────
def display(guitars):
    now = datetime.now().strftime("%Y-%m-%d  %H:%M:%S")
    print(f"\n{'─'*80}")
    print(f"  MULTI-SITE — All Guitars  >=excellent-              {now}")
    print(f"{'─'*80}")
    if not guitars:
        print("  (none passed the condition filter)")

    # Group by source for display
    by_source = {}
    for g in guitars:
        source = g.get("source", "unknown")
        by_source.setdefault(source, []).append(g)

    for source, source_guitars in by_source.items():
        print(f"\n  [{source}] {len(source_guitars)} guitars")
        for i, g in enumerate(source_guitars[:10], 1):  # Show first 10 per source
            brand, model = extract_brand_model(g["title"])
            year         = extract_year(g["title"])
            cond         = g.get("condition") or "unknown"
            hold = "  ON HOLD" if g.get("on_hold") else ""
            print(f"  {i:3}.  {brand:22} {model:24} {year:>8}  {g['price']:>10}  [{cond}]{hold}")

    print(f"{'─'*80}\n")


# ── main ─────────────────────────────────────────────────────────
def main():
    one_shot = "--once" in sys.argv

    sites = ["retrofret.com"]
    if WOODSTORE_ENABLED:
        sites.append("woodstore.fr")

    mode = "one-shot" if one_shot else "every 5 min"
    print(f"\n  searcher started — multi-site scraper ({mode})")
    print(f"  sites   → {', '.join(sites)}")
    print(f"  output  → {OUTPUT_FILE}")
    print(f"  cache   → {CACHE_FILE}\n")

    seen_ids        = load_seen_ids()
    cache           = load_cache()
    reverb_no_data  = load_reverb_no_data()
    sold_ids        = load_sold_ids()
    sold_candidates = load_sold_candidates()

    while True:
        print("  crawling …")
        all_guitars = scrape_all()
        print(f"  {len(all_guitars)} unique items found")

        # ── sold detection (full listing, before condition filter) ──
        if all_guitars:  # skip when scrape returns nothing (likely error)
            all_current_ids = {g["id"] for g in all_guitars}
            sold_candidates, newly_sold = check_sold(
                seen_ids, all_current_ids, sold_ids, sold_candidates
            )
            if newly_sold:
                mark_sold_batch(newly_sold)
                sold_ids.update(newly_sold)
            save_sold_candidates(sold_candidates)

        to_fetch = [g for g in all_guitars if g["id"] not in cache]
        if to_fetch:
            print(f"  fetching conditions for {len(to_fetch)} items …")
            with ThreadPoolExecutor(max_workers=10) as pool:
                conds = list(pool.map(fetch_condition, [g["url"] for g in to_fetch]))
            for g, c in zip(to_fetch, conds):
                cache[g["id"]] = c
            save_cache(cache)

        for g in all_guitars:
            g["condition"] = cache.get(g["id"])

        filtered = [g for g in all_guitars if condition_rank(g["condition"]) >= MIN_CONDITION_RANK]
        filtered.sort(key=lambda g: int(g["id"]), reverse=True)

        display(filtered)

        # flag any already-tracked guitars that are now on hold
        on_hold_now = {g["id"] for g in filtered if g.get("on_hold") and g["id"] in seen_ids}
        mark_on_hold(on_hold_now)

        # append only new guitars that are not on hold
        new_ids  = append_new([g for g in filtered if not g.get("on_hold")], seen_ids)
        seen_ids.update(new_ids)

        # retry any rows still missing Reverb prices
        reverb_no_data = backfill_reverb(reverb_no_data)

        if one_shot:
            print("  --once mode: done.")
            break

        print(f"  next crawl in 5 min …")
        time.sleep(INTERVAL)


if __name__ == "__main__":
    main()
