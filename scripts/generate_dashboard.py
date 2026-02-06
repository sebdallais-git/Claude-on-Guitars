#!/usr/bin/env python3
"""
generate_dashboard — creates dashboard.html and detail pages from listings.xlsx

Reads the Excel file and generates:
- dashboard.html (main overview)
- pages/recommendations.html (clickable top picks with details)
- pages/brand-{brand}.html (listings by brand)
"""

import os
import sys
import json
import re
from datetime import datetime
from collections import Counter, defaultdict

try:
    from openpyxl import load_workbook
except ImportError:
    print("Missing dep — run:  pip install openpyxl")
    sys.exit(1)

# Import currency conversion
sys.path.insert(0, os.path.dirname(__file__))
import currency

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.dirname(_SCRIPT_DIR)
_OUTPUTS = os.path.join(_PROJECT_ROOT, "outputs")
_PAGES = os.path.join(_PROJECT_ROOT, "pages")

XLSX_PATH = os.path.join(_OUTPUTS, "listings.xlsx")
DASHBOARD_PATH = os.path.join(_PROJECT_ROOT, "dashboard.html")

# Ensure pages directory exists
os.makedirs(_PAGES, exist_ok=True)


def parse_price(price_val):
    """Parse price from Excel - handle strings and numbers."""
    if not price_val:
        return None
    if isinstance(price_val, (int, float)):
        return float(price_val)
    if isinstance(price_val, str):
        try:
            return float(price_val.replace("$", "").replace(",", "").replace("€", "").strip())
        except (ValueError, AttributeError):
            return None
    return None


def load_listings():
    """Load all listings from Excel."""
    if not os.path.exists(XLSX_PATH):
        print(f"[!] {XLSX_PATH} not found")
        return []

    wb = load_workbook(XLSX_PATH, read_only=True)
    ws = wb.active

    # Check if we have Source column (new format) or not (old format)
    headers = [cell.value for cell in ws[1]]
    has_source = "Source" in headers

    listings = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        # Detect format based on headers
        if has_source:
            # New format with Source column
            if not row[10]:  # URL column
                continue
            listing = {
                "date": row[0],
                "source": row[1] or "retrofret.com",
                "brand": row[2] or "",
                "model": row[3] or "",
                "type": row[4] or "Unknown",
                "year": row[5] or "",
                "price": parse_price(row[6]),
                "reverb_low": parse_price(row[7]),
                "reverb_hi": parse_price(row[8]),
                "condition": row[9] or "unknown",
                "url": row[10] or "",
                "on_hold": row[11],
                "sold": row[12],
            }
        else:
            # Old format without Source column
            if not row[9]:  # URL column
                continue
            listing = {
                "date": row[0],
                "source": "retrofret.com",  # Default to retrofret
                "brand": row[1] or "",
                "model": row[2] or "",
                "type": row[3] or "Unknown",
                "year": row[4] or "",
                "price": parse_price(row[5]),
                "reverb_low": parse_price(row[6]),
                "reverb_hi": parse_price(row[7]),
                "condition": row[8] or "unknown",
                "url": row[9] or "",
                "on_hold": row[10],
                "sold": row[11],
            }

        listings.append(listing)

    wb.close()
    return listings


def compute_stats(listings):
    """Compute KPIs and stats from listings."""
    total = len(listings)
    active = sum(1 for l in listings if not l["on_hold"] and not l["sold"])
    on_hold = sum(1 for l in listings if l["on_hold"])
    sold = sum(1 for l in listings if l["sold"])

    # Parse prices - handle both numbers and strings
    prices = []
    for l in listings:
        p = l["price"]
        if p:
            if isinstance(p, (int, float)):
                prices.append(float(p))
            elif isinstance(p, str):
                # Remove $ and , then convert
                try:
                    prices.append(float(p.replace("$", "").replace(",", "").replace("€", "").strip()))
                except (ValueError, AttributeError):
                    pass

    avg_price = sum(prices) / len(prices) if prices else 0
    min_price = min(prices) if prices else 0
    max_price = max(prices) if prices else 0

    # Brands
    brands = Counter(l["brand"] for l in listings if l["brand"])
    top_brands = brands.most_common(8)

    # Types
    types = Counter(l["type"] for l in listings)

    # Decades (extract from year)
    def get_decade(year_str):
        match = re.search(r"(\d{4})", str(year_str))
        if match:
            year = int(match.group(1))
            return f"{year//10*10}s"
        return "Unknown"

    decades = Counter(get_decade(l["year"]) for l in listings)

    # Price buckets
    def price_bucket(price):
        if price < 1000:
            return "$0-1k"
        elif price < 2000:
            return "$1k-2k"
        elif price < 5000:
            return "$2k-5k"
        elif price < 10000:
            return "$5k-10k"
        else:
            return "$10k+"

    price_buckets = Counter(price_bucket(l["price"]) for l in listings if l["price"])

    return {
        "total": total,
        "active": active,
        "on_hold": on_hold,
        "sold": sold,
        "avg_price": avg_price,
        "min_price": min_price,
        "max_price": max_price,
        "brands": top_brands,
        "types": dict(types),
        "decades": dict(decades),
        "price_buckets": dict(price_buckets),
    }


def generate_recommendations_page(listings):
    """Generate pages/recommendations.html with top picks."""
    # For now, just show top 10 by price (you can integrate scorer.py later)
    active = [l for l in listings if not l["on_hold"] and not l["sold"] and l["price"]]
    top_picks = sorted(active, key=lambda x: x["price"], reverse=True)[:10]

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Top Recommendations — Vintage Guitar Collector</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Inter',sans-serif;background:#0f1117;color:#e2e4e9;padding:2rem}}
.wrap{{max-width:1000px;margin:0 auto}}
.header{{display:flex;justify-content:space-between;align-items:center;margin-bottom:2rem;padding-bottom:1rem;border-bottom:1px solid #1e2028}}
.title{{font-size:1.5rem;font-weight:700;color:#fff}}
.back{{color:#818cf8;text-decoration:none;font-size:.9rem;font-weight:500;padding:.5rem 1rem;background:#161822;border-radius:6px}}
.back:hover{{background:#1e2030}}
.listing{{background:#161822;border:1px solid #1e2028;border-radius:8px;padding:1.2rem;margin-bottom:1rem}}
.listing:hover{{border-color:#818cf8;background:#1a1c26}}
.l-head{{display:flex;justify-content:space-between;align-items:start;margin-bottom:.8rem}}
.l-title{{font-size:1.1rem;font-weight:600;color:#fff}}
.l-price{{font-size:1.2rem;font-weight:700;color:#4ade80}}
.l-meta{{display:flex;gap:1.5rem;font-size:.85rem;color:#6b7080;margin-bottom:.8rem}}
.l-meta span{{display:flex;align-items:center;gap:.3rem}}
.l-link{{display:inline-block;color:#818cf8;text-decoration:none;font-size:.85rem;font-weight:500;padding:.4rem .8rem;background:rgba(129,140,248,.1);border-radius:4px}}
.l-link:hover{{background:rgba(129,140,248,.2)}}
.badge{{display:inline-block;padding:.2rem .5rem;background:rgba(74,222,128,.1);color:#4ade80;font-size:.7rem;border-radius:4px;font-weight:600}}
</style>
</head>
<body>
<div class="wrap">
<div class="header">
  <div class="title">🎸 Top 10 Recommendations</div>
  <a href="../dashboard.html" class="back">← Back to Dashboard</a>
</div>
"""

    for i, listing in enumerate(top_picks, 1):
        # Determine currency and format with conversion
        curr = "EUR" if "woodstore" in listing["source"] else "USD"
        if listing['price']:
            price_str = currency.format_with_conversion(listing['price'], curr)
        else:
            price_str = "N/A"

        html += f"""
<div class="listing">
  <div class="l-head">
    <div class="l-title">#{i} — {listing['brand']} {listing['model']} ({listing['year']})</div>
    <div class="l-price">{price_str}</div>
  </div>
  <div class="l-meta">
    <span>📍 <strong>{listing['source']}</strong></span>
    <span>🎸 {listing['type']}</span>
    <span>✨ {listing['condition']}</span>
  </div>
  <a href="{listing['url']}" target="_blank" class="l-link">View Listing →</a>
</div>
"""

    html += """
</div>
</body>
</html>
"""

    with open(os.path.join(_PAGES, "recommendations.html"), "w") as f:
        f.write(html)

    print(f"  ✓ Generated pages/recommendations.html ({len(top_picks)} listings)")


def generate_brand_pages(listings):
    """Generate pages/brand-{brand}.html for each brand."""
    brands = defaultdict(list)
    for listing in listings:
        if listing["brand"]:
            brands[listing["brand"]].append(listing)

    for brand, brand_listings in brands.items():
        # Sort by price descending
        brand_listings = sorted(brand_listings, key=lambda x: x["price"] or 0, reverse=True)

        safe_brand = re.sub(r"[^a-z0-9]+", "-", brand.lower()).strip("-")
        filename = f"brand-{safe_brand}.html"

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{brand} Guitars — Vintage Guitar Collector</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Inter',sans-serif;background:#0f1117;color:#e2e4e9;padding:2rem}}
.wrap{{max-width:1200px;margin:0 auto}}
.header{{display:flex;justify-content:space-between;align-items:center;margin-bottom:2rem;padding-bottom:1rem;border-bottom:1px solid #1e2028}}
.title{{font-size:1.5rem;font-weight:700;color:#fff}}
.count{{font-size:1rem;color:#6b7080;font-weight:400}}
.back{{color:#818cf8;text-decoration:none;font-size:.9rem;font-weight:500;padding:.5rem 1rem;background:#161822;border-radius:6px}}
.back:hover{{background:#1e2030}}
.grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(320px,1fr));gap:1rem}}
.card{{background:#161822;border:1px solid #1e2028;border-radius:8px;padding:1rem;transition:all .2s}}
.card:hover{{border-color:#818cf8;transform:translateY(-2px)}}
.c-title{{font-size:1rem;font-weight:600;color:#fff;margin-bottom:.5rem}}
.c-meta{{font-size:.8rem;color:#6b7080;margin-bottom:.5rem}}
.c-price{{font-size:1.1rem;font-weight:700;color:#4ade80;margin-bottom:.5rem}}
.c-link{{display:inline-block;color:#818cf8;text-decoration:none;font-size:.8rem;font-weight:500}}
.c-link:hover{{text-decoration:underline}}
.badge{{display:inline-block;padding:.15rem .4rem;background:rgba(129,140,248,.1);color:#818cf8;font-size:.7rem;border-radius:3px;margin-top:.3rem}}
</style>
</head>
<body>
<div class="wrap">
<div class="header">
  <div>
    <div class="title">{brand} <span class="count">({len(brand_listings)} guitars)</span></div>
  </div>
  <a href="../dashboard.html" class="back">← Back to Dashboard</a>
</div>
<div class="grid">
"""

        for listing in brand_listings:
            # Determine currency and format with conversion
            curr = "EUR" if "woodstore" in listing["source"] else "USD"
            if listing['price']:
                price_str = currency.format_with_conversion(listing['price'], curr)
            else:
                price_str = "N/A"

            status = ""
            if listing["on_hold"]:
                status = '<span class="badge" style="background:rgba(251,191,36,.1);color:#fbbf24">ON HOLD</span>'
            elif listing["sold"]:
                status = '<span class="badge" style="background:rgba(248,113,113,.1);color:#f87171">SOLD</span>'

            html += f"""
<div class="card">
  <div class="c-title">{listing['model']}</div>
  <div class="c-meta">Year: {listing['year']} · {listing['type']} · {listing['condition']}</div>
  <div class="c-meta">Source: {listing['source']}</div>
  <div class="c-price">{price_str}</div>
  <a href="{listing['url']}" target="_blank" class="c-link">View Listing →</a>
  {status}
</div>
"""

        html += """
</div>
</div>
</body>
</html>
"""

        with open(os.path.join(_PAGES, filename), "w") as f:
            f.write(html)

    print(f"  ✓ Generated {len(brands)} brand pages")


def generate_dashboard(listings, stats):
    """Generate the main dashboard.html with real data."""
    # Read template and inject data
    # For simplicity, I'll create a complete dashboard here

    # Format data for JavaScript
    brands_js = [
        [b[0], b[1], ["#818cf8","#c084fc","#60a5fa","#4ade80","#fb923c","#f472b6","#fbbf24","#34d399"][i % 8]]
        for i, b in enumerate(stats["brands"])
    ]

    type_colors = {"Electric": "#818cf8", "Acoustic": "#4ade80", "Bass": "#60a5fa", "Unknown": "#4a4e5a"}
    types_data = [
        {"l": t, "n": n, "c": type_colors.get(t, "#4a4e5a")}  # Default to Unknown color
        for t, n in stats["types"].items()
    ]

    decade_list = ["1920s", "1930s", "1940s", "1950s", "1960s", "1970s", "1980s", "1990s", "2000s", "2010s", "2020s"]
    decades_data = [[d, stats["decades"].get(d, 0)] for d in decade_list]

    price_order = ["$0-1k", "$1k-2k", "$2k-5k", "$5k-10k", "$10k+"]
    prices_data = [[p, stats["price_buckets"].get(p, 0)] for p in price_order]

    # Top recommendations (simplified - just highest priced for now)
    active = [l for l in listings if not l["on_hold"] and not l["sold"] and l["price"]]
    top_recs = sorted(active, key=lambda x: x["price"], reverse=True)[:10]

    recs_js = []
    for i, rec in enumerate(top_recs, 1):
        recs_js.append({
            "r": i,
            "b": rec["brand"],
            "m": rec["model"],
            "y": rec["year"],
            "p": rec["price"],
            "s": 85,  # Placeholder score
            "sv": 100,
            "sa": 80,
            "sf": 50,
            "v1": int(rec["price"] * 1.1) if rec["price"] else None,
            "v2": int(rec["price"] * 1.2) if rec["price"] else None,
        })

    # Calculate budget info (TODO: read from budget.json)
    budget = 20000
    spent = 0

    # Generate HTML
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Vintage Guitar Collector — Dashboard</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
/* [CSS remains the same as original dashboard] */
*{{margin:0;padding:0;box-sizing:border-box}}
:root{{
  --bg:#0f1117;--card:#161822;--border:#1e2028;--card2:#1a1c24;
  --text:#e2e4e9;--muted:#6b7080;--dim:#4a4e5a;
  --indigo:#818cf8;--purple:#c084fc;
  --green:#4ade80;--amber:#fbbf24;--red:#f87171;
  --blue:#60a5fa;--orange:#fb923c;--pink:#f472b6;--teal:#34d399;
}}
body{{font-family:'Inter',sans-serif;background:var(--bg);color:var(--text);min-height:100vh;-webkit-font-smoothing:antialiased}}
header{{display:flex;justify-content:space-between;align-items:center;padding:1.2rem 2rem;border-bottom:1px solid var(--border);position:sticky;top:0;background:var(--bg);z-index:100}}
.logo{{font-size:.95rem;font-weight:600;color:#fff}}
.logo .g{{background:linear-gradient(135deg,var(--indigo),var(--purple));-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text}}
.gh{{color:var(--muted);text-decoration:none;font-size:.78rem;font-weight:500;display:flex;align-items:center;gap:.35rem}}
.gh:hover{{color:#fff}}
.gh svg{{width:15px;height:15px;fill:currentColor}}
.wrap{{max-width:1140px;margin:0 auto;padding:2rem 1.5rem 3rem}}
.sec{{font-size:.68rem;font-weight:600;text-transform:uppercase;letter-spacing:.12em;color:var(--dim);margin-bottom:.9rem;display:flex;align-items:center;gap:.7rem}}
.sec::after{{content:'';flex:1;height:1px;background:var(--border)}}
.kpis{{display:grid;grid-template-columns:repeat(auto-fit,minmax(148px,1fr));gap:.9rem;margin-bottom:2.4rem}}
.kpi{{background:var(--card);border:1px solid var(--border);border-radius:10px;padding:1.1rem 1rem}}
.kpi-l{{font-size:.67rem;color:var(--muted);font-weight:500;text-transform:uppercase;letter-spacing:.05em;margin-bottom:.4rem}}
.kpi-v{{font-size:1.55rem;font-weight:700;color:#fff;letter-spacing:-.02em}}
.kpi-s{{font-size:.68rem;color:var(--dim);margin-top:.25rem}}
.kpi.ci .kpi-v{{color:var(--indigo)}}.kpi.cg .kpi-v{{color:var(--green)}}.kpi.ca .kpi-v{{color:var(--amber)}}.kpi.cr .kpi-v{{color:var(--red)}}.kpi.cb .kpi-v{{color:var(--blue)}}.kpi.cp .kpi-v{{color:var(--purple)}}
.bp-bg{{background:var(--border);border-radius:3px;height:5px;margin-top:.55rem;overflow:hidden}}
.bp-fill{{height:100%;border-radius:3px;background:linear-gradient(90deg,var(--indigo),var(--purple))}}
.cg{{display:grid;grid-template-columns:1fr 1fr;gap:1rem;margin-bottom:2.4rem}}
@media(max-width:680px){{.cg{{grid-template-columns:1fr}}}}
.cc{{background:var(--card);border:1px solid var(--border);border-radius:10px;padding:1.3rem}}
.cc:hover{{border-color:var(--indigo);cursor:pointer}}
.ct{{font-size:.76rem;font-weight:600;color:var(--text);margin-bottom:.15rem}}
.cs{{font-size:.65rem;color:var(--dim);margin-bottom:.85rem}}
.bb{{display:flex;align-items:center;gap:.5rem;margin-bottom:.48rem;cursor:pointer;padding:.3rem;border-radius:4px;transition:background .2s}}
.bb:hover{{background:rgba(129,140,248,.05)}}
.bb-l{{font-size:.7rem;color:var(--muted);width:88px;text-align:right;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;flex-shrink:0}}
.bb-t{{flex:1;height:17px;background:var(--border);border-radius:3px;overflow:hidden}}
.bb-f{{height:100%;border-radius:3px;display:flex;align-items:center;padding-left:6px;font-size:.66rem;font-weight:600;color:rgba(255,255,255,.9)}}
.bb-n{{font-size:.66rem;color:var(--dim);width:20px;text-align:right;flex-shrink:0}}
.dn-w{{display:flex;align-items:center;gap:1.4rem}}
.dn-leg{{display:flex;flex-direction:column;gap:.5rem}}
.lg{{display:flex;align-items:center;gap:.45rem;font-size:.73rem;color:var(--muted)}}
.lg-d{{width:9px;height:9px;border-radius:50%;flex-shrink:0}}
.lg-n{{color:var(--text);font-weight:600;margin-left:auto;font-size:.72rem}}
.vb-w{{display:flex;align-items:flex-end;gap:5px;height:145px}}
.vb-c{{display:flex;flex-direction:column;align-items:center;flex:1;height:100%;justify-content:flex-end}}
.vb-r{{width:100%;border-radius:3px 3px 0 0;min-height:2px}}
.vb-l{{font-size:.58rem;color:var(--muted);margin-top:4px;text-align:center;white-space:nowrap}}
.vb-n{{font-size:.62rem;color:var(--muted);margin-bottom:2px;font-weight:500}}
.rc{{background:var(--card);border:1px solid var(--border);border-radius:10px;overflow:hidden;margin-bottom:2rem}}
.rc-s{{overflow-x:auto}}
.rc-s::-webkit-scrollbar{{height:3px}}
.rc-s::-webkit-scrollbar-thumb{{background:var(--border);border-radius:2px}}
table.rt{{width:100%;border-collapse:collapse}}
table.rt th{{font-size:.64rem;font-weight:600;color:var(--dim);text-transform:uppercase;letter-spacing:.05em;padding:.8rem .9rem;text-align:left;border-bottom:1px solid var(--border);background:#141620;white-space:nowrap}}
table.rt td{{font-size:.76rem;color:#c8cad0;padding:.68rem .9rem;border-bottom:1px solid #1a1c24;white-space:nowrap}}
table.rt tr:last-child td{{border-bottom:none}}
table.rt tbody tr:hover td{{background:rgba(129,140,248,.04);cursor:pointer}}
.sb{{display:inline-flex;align-items:center;justify-content:center;width:40px;height:23px;border-radius:5px;font-size:.71rem;font-weight:700}}
.sg{{background:rgba(74,222,128,.14);color:var(--green)}}.sy{{background:rgba(251,191,36,.14);color:var(--amber)}}.sr{{background:rgba(248,113,113,.14);color:var(--red)}}
.rk{{color:var(--indigo);font-weight:700}}
.pv{{color:var(--text);font-weight:600}}
.pn{{color:var(--dim);font-style:italic}}
.ds{{width:22px;height:5px;background:var(--border);border-radius:2px;overflow:hidden;display:inline-block}}
.df{{height:100%;border-radius:2px}}
footer{{border-top:1px solid var(--border);padding:1.2rem 2rem;display:flex;justify-content:space-between;font-size:.7rem;color:var(--dim)}}
.view-more{{display:inline-block;margin-top:1rem;color:var(--indigo);text-decoration:none;font-size:.8rem;font-weight:500}}
.view-more:hover{{text-decoration:underline}}
</style>
</head>
<body>

<header>
  <div class="logo">Vintage Guitar <span class="g">Collector</span></div>
  <a href="https://github.com/sebdallais-git" target="_blank" class="gh">
    <svg viewBox="0 0 24 24"><path d="M12 .297c-6.63 0-12 5.373-12 12 0 5.303 3.438 9.8 8.205 11.385.6.113.82-.258.82-.577 0-.285-.01-1.04-.015-2.04-3.338.724-4.042-1.61-4.042-1.61-.546-1.385-1.335-1.755-1.335-1.755-1.087-.744.084-.729.084-.729 1.205.084 1.838 1.236 1.838 1.236 1.07 1.835 2.809 1.305 3.495.998.108-.776.417-1.305.76-1.605-2.665-.3-5.466-1.332-5.466-5.93 0-1.31.465-2.38 1.235-3.22-.135-.303-.54-1.523.105-3.176 0 0 1.005-.322 3.3 1.23.96-.267 1.98-.399 3-.405 1.02.006 2.04.138 3 .405 2.28-1.552 3.285-1.23 3.285-1.23.645 1.653.24 2.873.12 3.176.765.84 1.23 1.91 1.23 3.22 0 4.61-2.805 5.625-5.475 5.92.42.36.81 1.096.81 2.22 0 1.606-.015 2.896-.015 3.286 0 .315.21.69.825.57C20.565 22.092 24 17.592 24 12.297c0-6.627-5.373-12-12-12"/></svg>
    sebdallais-git
  </a>
</header>

<div class="wrap">

<!-- ── workflow ── -->
<div class="sec">Agentic Workflow</div>
<div style="margin-bottom:2.4rem">
<svg viewBox="0 0 760 272" xmlns="http://www.w3.org/2000/svg">
  <defs>
    <marker id="ah" markerWidth="9" markerHeight="7" refX="8" refY="3.5" orient="auto" markerUnits="strokeWidth">
      <polygon points="0,0 9,3.5 0,7" fill="#4a4e5a"/>
    </marker>
    <linearGradient id="ag" x1="0%" y1="0%" x2="100%" y2="0%">
      <stop offset="0%" stop-color="#818cf8"/><stop offset="100%" stop-color="#c084fc"/>
    </linearGradient>
  </defs>

  <!-- ── ROW 1 ── -->
  <!-- Multi-site sources -->
  <rect x="4" y="10" width="116" height="64" rx="8" fill="#161822" stroke="#2a3a5c" stroke-width="1"/>
  <text x="62" y="33" text-anchor="middle" fill="#60a5fa" font-size="9.5" font-weight="600" font-family="Inter,sans-serif">retrofret.com</text>
  <text x="62" y="47" text-anchor="middle" fill="#4ade80" font-size="9.5" font-weight="600" font-family="Inter,sans-serif">woodstore.fr</text>
  <text x="62" y="61" text-anchor="middle" fill="#4a4e5a" font-size="8" font-family="Inter,sans-serif">sources</text>

  <!-- arrow -->
  <line x1="121" y1="42" x2="152" y2="42" stroke="#4a4e5a" stroke-width="1.5" marker-end="url(#ah)"/>

  <!-- Scraper -->
  <rect x="155" y="10" width="106" height="64" rx="8" fill="#161822" stroke="#282c3a" stroke-width="1"/>
  <rect x="155" y="10" width="106" height="3" rx="2" fill="url(#ag)"/>
  <text x="208" y="40" text-anchor="middle" fill="#e2e4e9" font-size="11" font-weight="600" font-family="Inter,sans-serif">Scraper</text>
  <text x="208" y="56" text-anchor="middle" fill="#4a4e5a" font-size="8.5" font-family="Inter,sans-serif">every 5 min</text>

  <!-- arrow -->
  <line x1="262" y1="42" x2="291" y2="42" stroke="#4a4e5a" stroke-width="1.5" marker-end="url(#ah)"/>

  <!-- listings.xlsx -->
  <rect x="294" y="10" width="118" height="64" rx="8" fill="#1a1c24" stroke="#2a2d3a" stroke-width="1"/>
  <text x="353" y="38" text-anchor="middle" fill="#e2e4e9" font-size="10.5" font-weight="600" font-family="Inter,sans-serif">listings.xlsx</text>
  <text x="353" y="55" text-anchor="middle" fill="#4a4e5a" font-size="8.5" font-family="Inter,sans-serif">{stats['total']} guitars</text>

  <!-- arrow right -->
  <line x1="413" y1="42" x2="449" y2="42" stroke="#4a4e5a" stroke-width="1.5" marker-end="url(#ah)"/>

  <!-- Watchdog -->
  <rect x="452" y="10" width="108" height="64" rx="8" fill="#161822" stroke="#282c3a" stroke-width="1"/>
  <rect x="452" y="10" width="108" height="3" rx="2" fill="url(#ag)"/>
  <text x="506" y="40" text-anchor="middle" fill="#e2e4e9" font-size="11" font-weight="600" font-family="Inter,sans-serif">Watchdog</text>
  <text x="506" y="56" text-anchor="middle" fill="#4a4e5a" font-size="8.5" font-family="Inter,sans-serif">monitors xlsx</text>

  <!-- arrow right -->
  <line x1="561" y1="42" x2="596" y2="42" stroke="#4a4e5a" stroke-width="1.5" marker-end="url(#ah)"/>

  <!-- Email + Telegram -->
  <rect x="599" y="10" width="140" height="64" rx="8" fill="#161822" stroke="#282c3a" stroke-width="1"/>
  <rect x="599" y="10" width="140" height="3" rx="2" fill="#4ade80"/>
  <text x="669" y="36" text-anchor="middle" fill="#e2e4e9" font-size="10" font-weight="600" font-family="Inter,sans-serif">Email</text>
  <text x="669" y="50" text-anchor="middle" fill="#4ade80" font-size="9" font-weight="500" font-family="Inter,sans-serif">+ Telegram</text>
  <text x="669" y="66" text-anchor="middle" fill="#4a4e5a" font-size="7.5" font-family="Inter,sans-serif">notifications</text>

  <!-- ── vertical arrow: listings → Scorer ── -->
  <line x1="353" y1="75" x2="353" y2="148" stroke="#4a4e5a" stroke-width="1.5" marker-end="url(#ah)" stroke-dasharray="4,3"/>

  <!-- ── ROW 2 ── -->
  <!-- collection.json -->
  <rect x="4" y="152" width="120" height="64" rx="8" fill="#161822" stroke="#2a3a5c" stroke-width="1"/>
  <text x="64" y="180" text-anchor="middle" fill="#60a5fa" font-size="10" font-weight="600" font-family="Inter,sans-serif">collection.json</text>
  <text x="64" y="197" text-anchor="middle" fill="#4a4e5a" font-size="8.5" font-family="Inter,sans-serif">your guitars</text>

  <!-- arrow -->
  <line x1="125" y1="184" x2="153" y2="184" stroke="#4a4e5a" stroke-width="1.5" marker-end="url(#ah)"/>

  <!-- Valuation -->
  <rect x="156" y="152" width="108" height="64" rx="8" fill="#161822" stroke="#282c3a" stroke-width="1"/>
  <rect x="156" y="152" width="108" height="3" rx="2" fill="url(#ag)"/>
  <text x="210" y="182" text-anchor="middle" fill="#e2e4e9" font-size="11" font-weight="600" font-family="Inter,sans-serif">Valuation</text>
  <text x="210" y="198" text-anchor="middle" fill="#4a4e5a" font-size="8.5" font-family="Inter,sans-serif">Reverb API</text>

  <!-- arrow -->
  <line x1="265" y1="184" x2="293" y2="184" stroke="#4a4e5a" stroke-width="1.5" marker-end="url(#ah)"/>

  <!-- Scorer -->
  <rect x="296" y="152" width="114" height="64" rx="8" fill="#161822" stroke="#282c3a" stroke-width="1"/>
  <rect x="296" y="152" width="114" height="3" rx="2" fill="url(#ag)"/>
  <text x="353" y="182" text-anchor="middle" fill="#e2e4e9" font-size="11" font-weight="600" font-family="Inter,sans-serif">Scorer</text>
  <text x="353" y="198" text-anchor="middle" fill="#4a4e5a" font-size="8.5" font-family="Inter,sans-serif">3-dim scoring</text>

  <!-- arrow right -->
  <line x1="411" y1="184" x2="446" y2="184" stroke="#4a4e5a" stroke-width="1.5" marker-end="url(#ah)"/>

  <!-- Recommendations (clickable) -->
  <a href="pages/recommendations.html" style="cursor:pointer">
    <rect x="449" y="152" width="152" height="64" rx="8" fill="#161822" stroke="#282c3a" stroke-width="1"/>
    <rect x="449" y="152" width="152" height="3" rx="2" fill="#4ade80"/>
    <text x="525" y="181" text-anchor="middle" fill="#e2e4e9" font-size="10.5" font-weight="600" font-family="Inter,sans-serif">Recommendations</text>
    <text x="525" y="198" text-anchor="middle" fill="#4ade80" font-size="8.5" font-family="Inter,sans-serif" style="text-decoration:underline">top 10 picks</text>
  </a>

  <!-- ── budget.json (small box) ── -->
  <rect x="4" y="234" width="120" height="34" rx="7" fill="#161822" stroke="#2a3a5c" stroke-width="1"/>
  <text x="64" y="256" text-anchor="middle" fill="#60a5fa" font-size="9.5" font-weight="600" font-family="Inter,sans-serif">budget.json  ·  ${budget/1000:.0f}k</text>

  <!-- L-shaped arrow: budget → Scorer bottom -->
  <path d="M 125,251 L 351,251 L 351,217" fill="none" stroke="#4a4e5a" stroke-width="1.5" marker-end="url(#ah)" stroke-dasharray="4,3"/>

  <!-- ── legend row ── -->
  <!-- source dot -->
  <rect x="4" y="265" width="8" height="8" rx="2" fill="#2a3a5c"/>
  <text x="16" y="273" fill="#4a4e5a" font-size="7.5" font-family="Inter,sans-serif">data source</text>
  <!-- agent dot -->
  <rect x="90" y="265" width="8" height="8" rx="2" fill="url(#ag)"/>
  <text x="102" y="273" fill="#4a4e5a" font-size="7.5" font-family="Inter,sans-serif">agent</text>
  <!-- output dot -->
  <rect x="158" y="265" width="8" height="8" rx="2" fill="#4ade80"/>
  <text x="170" y="273" fill="#4a4e5a" font-size="7.5" font-family="Inter,sans-serif">output</text>
  <!-- dashed -->
  <line x1="222" y1="269" x2="258" y2="269" stroke="#4a4e5a" stroke-width="1.5" stroke-dasharray="4,3"/>
  <text x="264" y="273" fill="#4a4e5a" font-size="7.5" font-family="Inter,sans-serif">data flow</text>
</svg>
</div>

<div class="sec">At a Glance</div>
<div class="kpis">
  <div class="kpi ci">
    <div class="kpi-l">Total Tracked</div>
    <div class="kpi-v">{stats['total']}</div>
    <div class="kpi-s">all sites</div>
  </div>
  <div class="kpi cg">
    <div class="kpi-l">Active</div>
    <div class="kpi-v">{stats['active']}</div>
    <div class="kpi-s">available now</div>
  </div>
  <div class="kpi ca">
    <div class="kpi-l">On Hold</div>
    <div class="kpi-v">{stats['on_hold']}</div>
    <div class="kpi-s">seller holding</div>
  </div>
  <div class="kpi cr">
    <div class="kpi-l">Sold</div>
    <div class="kpi-v">{stats['sold']}</div>
    <div class="kpi-s">confirmed sold</div>
  </div>
  <div class="kpi cb">
    <div class="kpi-l">Avg Price</div>
    <div class="kpi-v">${stats['avg_price']/1000:.1f}k</div>
    <div class="kpi-s">range ${stats['min_price']:.0f} – ${stats['max_price']/1000:.0f}k</div>
  </div>
  <div class="kpi cp">
    <div class="kpi-l">Budget</div>
    <div class="kpi-v">${budget/1000:.0f}k</div>
    <div class="kpi-s">${spent:.0f} spent · ${(budget-spent)/1000:.0f}k left</div>
    <div class="bp-bg"><div class="bp-fill" style="width:{spent/budget*100:.0f}%"></div></div>
  </div>
</div>

<div class="sec">Market Breakdown</div>
<div class="cg">
  <div class="cc">
    <div class="ct">Top Brands</div>
    <div class="cs">by listing count (click to view)</div>
    <div id="brands"></div>
  </div>
  <div class="cc">
    <div class="ct">Guitar Types</div>
    <div class="cs">acoustic · electric · bass</div>
    <div class="dn-w">
      <svg id="donut" width="120" height="120" viewBox="0 0 120 120"></svg>
      <div class="dn-leg" id="dn-leg"></div>
    </div>
  </div>
  <div class="cc">
    <div class="ct">By Decade</div>
    <div class="cs">manufacturing year</div>
    <div class="vb-w" id="decades"></div>
  </div>
  <div class="cc">
    <div class="ct">Price Distribution</div>
    <div class="cs">listing price in USD</div>
    <div class="vb-w" id="prices"></div>
  </div>
</div>

<div class="sec">Top 10 Recommendations</div>
<div class="rc"><div class="rc-s">
  <table class="rt">
    <thead><tr>
      <th>#</th>
      <th>Brand</th>
      <th>Model</th>
      <th>Year</th>
      <th style="text-align:right">Price</th>
      <th>Score</th>
      <th>Val · App · Fit</th>
      <th style="text-align:right">+1 Year</th>
      <th style="text-align:right">+2 Years</th>
    </tr></thead>
    <tbody id="recs"></tbody>
  </table>
</div></div>
<a href="pages/recommendations.html" class="view-more">View Full Details →</a>

</div>

<footer>
  <span>&copy; 2026 Sebastien DALLAIS</span>
  <span>vintage-guitar-collector · updated {datetime.now().strftime("%Y-%m-%d %H:%M")}</span>
</footer>

<script>
// Data generated from listings.xlsx
const BRANDS = {json.dumps(brands_js)};
const TYPES = {json.dumps(types_data)};
const DECADES = {json.dumps(decades_data)};
const PRICES = {json.dumps(prices_data)};
const RECS = {json.dumps(recs_js)};

// Brands with click handlers
(function(){{
  const max = BRANDS[0][1];
  document.getElementById('brands').innerHTML = BRANDS.map(([name,n,c])=>{{
    const safeName = name.toLowerCase().replace(/[^a-z0-9]+/g, '-').replace(/^-|-$/g, '');
    return `<div class="bb" onclick="window.location='pages/brand-${{safeName}}.html'">
       <div class="bb-l">${{name}}</div>
       <div class="bb-t"><div class="bb-f" style="width:${{n/max*100}}%;background:${{c}}">${{n}}</div></div>
       <div class="bb-n">${{n}}</div>
     </div>`;
  }}).join('');
}})();

// Donut chart
(function(){{
  const total = TYPES.reduce((s,t)=>s+t.n,0);
  const cx=60, cy=60, R=48, r=30;
  let angle = -90, paths = '';
  TYPES.forEach(t=>{{
    const sweep = (t.n/total)*360;
    const s = angle*Math.PI/180, e = (angle+sweep)*Math.PI/180;
    const lg = sweep > 180 ? 1 : 0;
    const ox1=cx+R*Math.cos(s), oy1=cy+R*Math.sin(s);
    const ox2=cx+R*Math.cos(e), oy2=cy+R*Math.sin(e);
    const ix1=cx+r*Math.cos(e), iy1=cy+r*Math.sin(e);
    const ix2=cx+r*Math.cos(s), iy2=cy+r*Math.sin(s);
    paths += `<path d="M${{ox1}},${{oy1}} A${{R}},${{R}} 0 ${{lg}},1 ${{ox2}},${{oy2}} L${{ix1}},${{iy1}} A${{r}},${{r}} 0 ${{lg}},0 ${{ix2}},${{iy2}}Z" fill="${{t.c}}" opacity=".85"/>`;
    angle += sweep;
  }});
  paths += `<text x="60" y="56" text-anchor="middle" fill="#e2e4e9" font-size="15" font-weight="700" font-family="Inter,sans-serif">${{total}}</text>`;
  paths += `<text x="60" y="72" text-anchor="middle" fill="#4a4e5a" font-size="8" font-family="Inter,sans-serif">total</text>`;
  document.getElementById('donut').innerHTML = paths;
  document.getElementById('dn-leg').innerHTML = TYPES.map(t=>
    `<div class="lg"><div class="lg-d" style="background:${{t.c}}"></div>${{t.l}}<span class="lg-n">${{t.n}}</span></div>`
  ).join('');
}})();

// Vertical bars
function vbars(id, data, colorFn){{
  const max = Math.max(...data.map(d=>d[1]));
  document.getElementById(id).innerHTML = data.map(([label,val])=>{{
    const h = val > 0 ? Math.max(8, (val/max)*128) : 2;
    return `<div class="vb-c">
              <div class="vb-n">${{val}}</div>
              <div class="vb-r" style="height:${{h}}px;background:${{colorFn(val,max)}}"></div>
              <div class="vb-l">${{label}}</div>
            </div>`;
  }}).join('');
}}
vbars('decades', DECADES, (v,m)=>`rgba(129,140,248,${{.3+v/m*.7}})`);
vbars('prices',  PRICES,  (v,m)=>`rgba(192,132,252,${{.3+v/m*.7}})`);

// Recommendations with click to detail page
(function(){{
  const f$ = n => n!=null ? '$'+Number(n).toLocaleString() : 'N/A';
  function slot(pct,color){{
    return `<span class="ds"><span class="df" style="width:${{pct}}%;background:${{color}}"></span></span>`;
  }}
  document.getElementById('recs').innerHTML = RECS.map(r=>{{
    const sc = r.s>=80?'sg':r.s>=60?'sy':'sr';
    return `<tr onclick="window.location='pages/recommendations.html'">
      <td class="rk">${{r.r}}</td>
      <td>${{r.b}}</td>
      <td style="color:var(--muted)">${{r.m}}</td>
      <td>${{r.y}}</td>
      <td style="text-align:right" class="${{r.p!=null?'pv':'pn'}}">${{f$(r.p)}}</td>
      <td><span class="sb ${{sc}}">${{r.s}}</span></td>
      <td style="display:flex;gap:3px;align-items:center">
        ${{slot(r.sv,'#818cf8')}}${{slot(r.sa,'#4ade80')}}${{slot(r.sf,'#60a5fa')}}
      </td>
      <td class="pv" style="text-align:right">${{f$(r.v1)}}</td>
      <td class="pv" style="text-align:right">${{f$(r.v2)}}</td>
    </tr>`;
  }}).join('');
}})();
</script>
</body>
</html>
"""

    with open(DASHBOARD_PATH, "w") as f:
        f.write(html)

    print(f"  ✓ Generated dashboard.html")


def main():
    print("\n=== Generating Dashboard ===\n")

    listings = load_listings()
    if not listings:
        print("[!] No listings found in Excel")
        return

    print(f"  Loaded {len(listings)} listings")

    stats = compute_stats(listings)

    generate_dashboard(listings, stats)
    generate_recommendations_page(listings)
    generate_brand_pages(listings)

    print("\n✅ Dashboard generation complete!")
    print(f"   • dashboard.html")
    print(f"   • pages/recommendations.html")
    print(f"   • pages/brand-*.html ({len(stats['brands'])} brands)\n")


if __name__ == "__main__":
    main()
