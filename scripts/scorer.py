#!/usr/bin/env python3
"""
scorer — scores active listings and writes top recommendations.

Scoring dimensions (weights tunable in data/budget.json):
  1. Value opportunity   — how far below (or above) the Reverb market range
  2. Appreciation        — projected annual rate, with golden era boost
  3. Collection fit      — diversification bonus, duplicate penalty, iconic model boost
  4. Condition           — guitar condition mapped to 0-100
  5. Iconic status       — how many top-100 guitarists played this model (rank-weighted)

Backward-compatible: dimensions only activate if their weight key exists
in budget.json (e.g. 'condition', 'iconic').

Run standalone:
    python3 scorer.py
"""

import json
import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from valuation import appreciation_rate, project_value, read_collection, _parse_year

# ML inference — optional, graceful fallback when not installed or no models
try:
    import ml_predict as _ml_predict
except ImportError:
    _ml_predict = None

# ── paths ─────────────────────────────────────────────────────────
_SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.dirname(_SCRIPT_DIR)
_DATA         = os.path.join(_PROJECT_ROOT, "data")
_OUTPUTS      = os.path.join(_PROJECT_ROOT, "outputs")
_KNOWLEDGE    = os.path.join(_DATA, "knowledge")

XLSX_PATH   = os.path.join(_OUTPUTS, "listings.xlsx")
BUDGET_FILE = os.path.join(_DATA, "budget.json")


_DEFAULT_BUDGET = {
    "total": 20000,
    "spent": 0,
    "weights": {
        "value": 0.30,
        "appreciate": 0.25,
        "fit": 0.25,
        "condition": 0.20,
    },
    "top_n": 10,
}


def _load_budget():
    if not os.path.exists(BUDGET_FILE):
        print(f"  Warning: {BUDGET_FILE} not found, using defaults")
        return dict(_DEFAULT_BUDGET)
    with open(BUDGET_FILE) as f:
        return json.load(f)


# ── condition scores ──────────────────────────────────────────────
_CONDITION_SCORES = {
    "mint":          100,
    "near mint":      95,
    "excellent+":     90,
    "excellent":      85,
    "excellent-":     80,
    "very good+":     70,
    "very good":      60,
    "very good-":     50,
    "good+":          40,
    "good":           30,
    "good-":          20,
    "fair":           10,
    "poor":            0,
}


def _score_condition(condition_str):
    """Map a condition string to 0-100. Returns 50 for unknown conditions."""
    if not condition_str:
        return 50.0
    normalized = condition_str.strip().lower()
    # Try exact match first
    if normalized in _CONDITION_SCORES:
        return float(_CONDITION_SCORES[normalized])
    # Try partial match (e.g. "Excellent Plus" → "excellent+")
    for key, score in _CONDITION_SCORES.items():
        if key.replace("+", " plus").replace("-", " minus") in normalized:
            return float(score)
        if normalized in key or key in normalized:
            return float(score)
    return 50.0


# ── iconic models knowledge base ─────────────────────────────────
_iconic_cache = None


def _load_iconic_models():
    """Load iconic models from data/knowledge/iconic_models.json."""
    global _iconic_cache
    if _iconic_cache is not None:
        return _iconic_cache

    _iconic_cache = []
    path = os.path.join(_KNOWLEDGE, "iconic_models.json")
    if os.path.exists(path):
        try:
            with open(path) as f:
                data = json.load(f)
            _iconic_cache = data.get("models", [])
        except (json.JSONDecodeError, KeyError):
            pass

    return _iconic_cache


def _match_iconic_model(brand, model):
    """Find the best matching iconic model entry (longest model match wins).

    Returns the iconic model dict or None.
    """
    iconic = _load_iconic_models()
    if not iconic:
        return None

    brand_l = brand.lower()
    model_l = model.lower()
    best = None
    best_len = 0

    for entry in iconic:
        entry_brand = entry["brand"].lower()
        entry_model = entry["model"].lower()

        # Brand must match (any word overlap)
        brand_words = set(re.split(r"[^a-z]+", brand_l)) - {""}
        iconic_words = set(re.split(r"[^a-z]+", entry_brand)) - {""}
        if not (brand_words & iconic_words):
            continue

        # Model must be a substring match (in either direction)
        if entry_model in model_l or model_l in entry_model:
            match_len = len(entry_model)
            if match_len > best_len:
                best = entry
                best_len = match_len

    return best


# ── top guitarists knowledge base (iconic scoring) ───────────────
_guitarists_cache = None
_iconic_score_cache = None


def _load_top_guitarists():
    """Load top guitarists from data/knowledge/top_guitarists.json."""
    global _guitarists_cache
    if _guitarists_cache is not None:
        return _guitarists_cache
    _guitarists_cache = []
    path = os.path.join(_KNOWLEDGE, "top_guitarists.json")
    if os.path.exists(path):
        try:
            with open(path) as f:
                data = json.load(f)
            _guitarists_cache = data.get("guitarists", [])
        except (json.JSONDecodeError, KeyError):
            pass
    return _guitarists_cache


def _build_iconic_scores():
    """Build rank-weighted association counts per model, return (lookup, max_score).

    Each guitarist contributes (101 - rank) / 100 points, so #1 = 1.0, #100 = 0.01.
    Lookup key: "brand_lower|model_lower".
    """
    global _iconic_score_cache
    if _iconic_score_cache is not None:
        return _iconic_score_cache

    guitarists = _load_top_guitarists()
    model_scores = {}  # "brand|model" → weighted count

    for g in guitarists:
        weight = (101 - g["rank"]) / 100.0
        for brand, model in g.get("guitars", []):
            key = f"{brand.lower()}|{model.lower()}"
            model_scores[key] = model_scores.get(key, 0.0) + weight

    max_score = max(model_scores.values()) if model_scores else 1.0
    _iconic_score_cache = (model_scores, max_score)
    return _iconic_score_cache


def _score_iconic(brand, model):
    """0-100: how many top-100 guitarists are associated with this model.

    Weighted by rank — higher-ranked guitarists contribute more.
    Score = (weighted_count / max_weighted_count) * 100, capped at 100.
    Uses the same substring matching logic as _match_iconic_model().
    """
    guitarists = _load_top_guitarists()
    if not guitarists:
        return 50.0  # neutral when no data

    model_scores, max_score = _build_iconic_scores()
    brand_l = brand.lower()
    model_l = model.lower()

    # Try exact key match first, then substring matching
    key = f"{brand_l}|{model_l}"
    if key in model_scores:
        return min(100.0, model_scores[key] / max_score * 100)

    # Substring matching: same logic as _match_iconic_model
    best_score = 0.0
    best_len = 0
    for mkey, wscore in model_scores.items():
        k_brand, k_model = mkey.split("|", 1)

        # Brand must share a word
        brand_words = set(re.split(r"[^a-z]+", brand_l)) - {""}
        key_words = set(re.split(r"[^a-z]+", k_brand)) - {""}
        if not (brand_words & key_words):
            continue

        # Model substring match (either direction), longest wins
        if k_model in model_l or model_l in k_model:
            match_len = len(k_model)
            if match_len > best_len:
                best_score = wscore
                best_len = match_len

    if best_len > 0:
        return min(100.0, best_score / max_score * 100)
    return 0.0


# ── read active listings ──────────────────────────────────────────
def read_active_listings():
    """Return listings from listings.xlsx that are not sold and not on hold.

    Column layout (13 cols, 0-indexed):
      0: Date Arrived  1: Source       2: Brand/Make   3: Model
      4: Type          5: Year         6: Price        7: Reverb Low
      8: Reverb High   9: Condition   10: URL         11: On Hold
     12: Sold Date
    """
    if not os.path.exists(XLSX_PATH):
        return []
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb.active
    listings = []
    for row in ws.iter_rows(min_row=2, max_col=13, values_only=True):
        url = str(row[10]) if row[10] else ""          # col 11 = URL
        # Support both retrofret ProductID and woodstore/other slug URLs
        m = re.search(r"ProductID=(\d+)", url)
        if m:
            pid = m.group(1)
        else:
            m2 = re.search(r"/products?/([^/?]+)", url)
            if not m2:
                continue
            pid = m2.group(1)
        if row[11] is not None or row[12] is not None:  # On Hold or Sold Date
            continue
        price = None
        if row[6] is not None:                           # col 7 = Price
            try:
                price = float(row[6])
            except (ValueError, TypeError):
                pass
        listings.append({
            "id":         pid,
            "brand":      str(row[2] or ""),             # col 3 = Brand
            "model":      str(row[3] or ""),             # col 4 = Model
            "type":       str(row[4] or ""),             # col 5 = Type
            "year":       str(row[5] or ""),             # col 6 = Year
            "price":      price,
            "reverb_lo":  row[7],                        # col 8 = Reverb Low
            "reverb_hi":  row[8],                        # col 9 = Reverb High
            "condition":  str(row[9] or ""),             # col 10 = Condition
            "url":        url,
        })
    wb.close()
    return listings


# ── scoring ───────────────────────────────────────────────────────
def _score_value(price, reverb_lo, reverb_hi):
    """
    0–100: how good a deal is this listing vs the Reverb market range.
    At or below reverb_lo → 100.  At reverb_hi → 50.  Above → below 50.
    Returns 50 (neutral) when price or Reverb data is missing.
    """
    if price is None or reverb_lo is None or reverb_hi is None:
        return 50.0
    mid = (reverb_lo + reverb_hi) / 2
    if price <= reverb_lo:
        return 100.0
    if price <= mid:
        return 100.0 - 25.0 * (price - reverb_lo) / (mid - reverb_lo)
    if price <= reverb_hi:
        return 75.0 - 25.0 * (price - mid) / (reverb_hi - mid)
    # above reverb_hi: linear drop toward 0
    overshoot = (price - reverb_hi) / reverb_hi
    return max(0.0, 50.0 - 50.0 * overshoot)


def _score_appreciation(brand, year, model=None):
    """0–100: maps the 0–12% annual rate range linearly onto 0–100.

    Adds a +20 golden era boost when the listing year falls within an
    iconic model's golden era.
    """
    rate = appreciation_rate(brand, year, model=model)
    score = min(100.0, rate / 0.12 * 100)

    # Golden era boost from iconic models knowledge base
    iconic = _match_iconic_model(brand, model or "")
    if iconic:
        yr = _parse_year(year)
        era = iconic.get("golden_era")
        if yr and era and era[0] <= yr <= era[1]:
            score = min(100.0, score + 20)

    return score


def _score_fit(entry, collection):
    """
    0–100 collection-fit score.  Starts at 50.
    +20 brand not yet owned (diversification).
    +15 type is under-represented (< 25 % of collection).
    -25 exact brand+model already owned.
    +N  popularity boost from iconic model (0-20 points).
    """
    score = 50
    if not collection:
        # No collection to compare; still apply iconic boost
        iconic = _match_iconic_model(entry["brand"], entry["model"])
        if iconic:
            score += iconic.get("boost", 0)
        return max(0, min(100, score))

    brands_owned = {g["brand"].lower() for g in collection}
    models_owned = {(g["brand"].lower(), g["model"].lower()) for g in collection}
    brand_l      = entry["brand"].lower()
    model_l      = entry["model"].lower()
    type_l       = entry.get("type", "").lower()

    if (brand_l, model_l) in models_owned:
        score -= 25
    elif brand_l not in brands_owned:
        score += 20

    if type_l:
        type_count = sum(1 for g in collection if g["type"].lower() == type_l)
        if type_count / len(collection) < 0.25:
            score += 15

    # Iconic model popularity boost
    iconic = _match_iconic_model(entry["brand"], entry["model"])
    if iconic:
        score += iconic.get("boost", 0)

    return max(0, min(100, score))


def score_listing(entry, collection, budget):
    """
    Composite 0–100 score + per-dimension breakdown.
    Supports 3/4/5-dim scoring based on which weight keys exist in budget.json.
    When ml_enabled is True in budget, blends ML scores using ml_blend factor.
    Returns (total, breakdown_dict).
    """
    w = budget["weights"]
    s_val  = _score_value(entry["price"], entry.get("reverb_lo"), entry.get("reverb_hi"))
    s_app  = _score_appreciation(entry["brand"], entry["year"], model=entry.get("model"))
    s_fit  = _score_fit(entry, collection)
    s_cond = _score_condition(entry.get("condition", ""))
    s_icon = _score_iconic(entry["brand"], entry.get("model", ""))

    # Build rule-based total from whichever dimensions have weights
    rule_total = w["value"] * s_val + w["appreciate"] * s_app + w["fit"] * s_fit
    if "condition" in w:
        rule_total += w["condition"] * s_cond
    if "iconic" in w:
        rule_total += w["iconic"] * s_icon

    breakdown = {
        "value":        round(s_val, 1),
        "appreciation": round(s_app, 1),
        "fit":          round(s_fit, 1),
        "condition":    round(s_cond, 1),
        "iconic":       round(s_icon, 1),
        "ml_total":     None,
        "ml_buy_prob":  None,
        "ml_price":     None,
        "rule_total":   round(rule_total, 1),
    }

    # Hybrid ML blending (when enabled and models available)
    total = rule_total
    if budget.get("ml_enabled") and _ml_predict and _ml_predict.is_ml_available():
        ml_blend = budget.get("ml_blend", 0.3)
        ml_result = _ml_predict.ml_score_listing(entry, collection, budget, breakdown)
        if ml_result:
            if ml_result.get("ml_total") is not None:
                total = (1 - ml_blend) * rule_total + ml_blend * ml_result["ml_total"]
            breakdown["ml_total"] = ml_result.get("ml_total")
            breakdown["ml_buy_prob"] = ml_result.get("ml_buy_prob")
            breakdown["ml_price"] = ml_result.get("ml_price")

    return round(total, 1), breakdown


# ── write recommendations sheet ───────────────────────────────────
REC_HEADERS = [
    "Rank", "Brand", "Model", "Type", "Year", "Price ($)",
    "Reverb Low $", "Reverb High $", "Condition", "Cond.",
    "Score", "Value", "Apprec.", "Fit", "Iconic",
    "Value in 1 Year ($)", "Value in 2 Years ($)",
    "URL",
    "ML Score", "Buy Prob", "ML Price",
]
REC_WIDTHS = [6, 22, 26, 18, 10, 14, 14, 14, 14, 8, 8, 8, 10, 6, 8, 20, 22, 65, 10, 10, 14]

HEADER_FILL  = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
URL_FONT     = Font(color="0563C1", underline="single", size=10)
CURRENCY_FMT = "$#,##0.00"


def _score_fill(score):
    """Green >= 80, yellow >= 60, red below."""
    if score >= 80:
        return PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    if score >= 60:
        return PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    return PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def write_recommendations(scored, budget):
    """Write (or replace) the 'Recommendations' sheet in listings.xlsx."""
    wb = load_workbook(XLSX_PATH)

    if "Recommendations" in wb.sheetnames:
        del wb["Recommendations"]
    ws = wb.create_sheet("Recommendations")

    # header row
    for col, (header, width) in enumerate(zip(REC_HEADERS, REC_WIDTHS), 1):
        cell           = ws.cell(row=1, column=col, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 20

    budget_remaining = budget["total"] - budget["spent"]

    for rank, (entry, total, breakdown) in enumerate(scored, 1):
        row   = rank + 1
        price = entry["price"] or 0
        over  = price > budget_remaining

        ws.cell(row=row, column=1,  value=rank)
        ws.cell(row=row, column=2,  value=entry["brand"])
        ws.cell(row=row, column=3,  value=entry["model"])
        ws.cell(row=row, column=4,  value=entry["type"])
        ws.cell(row=row, column=5,  value=entry["year"])

        if price:
            ws.cell(row=row, column=6, value=price)
            ws.cell(row=row, column=6).number_format = CURRENCY_FMT
        if entry.get("reverb_lo"):
            ws.cell(row=row, column=7, value=entry["reverb_lo"])
            ws.cell(row=row, column=7).number_format = CURRENCY_FMT
        if entry.get("reverb_hi"):
            ws.cell(row=row, column=8, value=entry["reverb_hi"])
            ws.cell(row=row, column=8).number_format = CURRENCY_FMT
        ws.cell(row=row, column=9, value=entry["condition"])

        # Condition sub-score (new "Cond." column)
        ws.cell(row=row, column=10, value=breakdown["condition"])

        # composite score with colour
        score_cell      = ws.cell(row=row, column=11, value=total)
        score_cell.fill = _score_fill(total)

        ws.cell(row=row, column=12, value=breakdown["value"])
        ws.cell(row=row, column=13, value=breakdown["appreciation"])
        ws.cell(row=row, column=14, value=breakdown["fit"])
        ws.cell(row=row, column=15, value=breakdown["iconic"])

        # projected values (Reverb midpoint, or listing price as fallback)
        if entry.get("reverb_lo") and entry.get("reverb_hi"):
            base = (entry["reverb_lo"] + entry["reverb_hi"]) / 2
        else:
            base = price or None
        v1y = project_value(base, entry["brand"], entry["year"], 1,
                            model=entry.get("model"))
        v2y = project_value(base, entry["brand"], entry["year"], 2,
                            model=entry.get("model"))
        if v1y:
            ws.cell(row=row, column=16, value=v1y)
            ws.cell(row=row, column=16).number_format = CURRENCY_FMT
        if v2y:
            ws.cell(row=row, column=17, value=v2y)
            ws.cell(row=row, column=17).number_format = CURRENCY_FMT

        # URL
        url_cell           = ws.cell(row=row, column=18, value=entry["url"])
        url_cell.hyperlink = entry["url"]
        url_cell.font      = URL_FONT

        # ML columns (19-21) — only populated when ML is active
        if breakdown.get("ml_total") is not None:
            ws.cell(row=row, column=19, value=breakdown["ml_total"])
        if breakdown.get("ml_buy_prob") is not None:
            ws.cell(row=row, column=20, value=breakdown["ml_buy_prob"])
        if breakdown.get("ml_price") is not None:
            ws.cell(row=row, column=21, value=breakdown["ml_price"])
            ws.cell(row=row, column=21).number_format = CURRENCY_FMT

        # grey out the whole row when over remaining budget
        if over:
            grey = Font(color="808080", size=10)
            for col in range(1, 22):
                ws.cell(row=row, column=col).font = grey
            url_cell.font = Font(color="808080", size=10, underline="single")

    wb.save(XLSX_PATH)


# ── main ──────────────────────────────────────────────────────────
def run():
    """Score all active listings, print top N, write Recommendations sheet."""
    budget     = _load_budget()
    collection = read_collection()
    listings   = read_active_listings()

    if not listings:
        print("  No active listings found in listings.xlsx.")
        return

    budget_remaining = budget["total"] - budget["spent"]
    w = budget["weights"]
    dims = f"{len(w)}-dim"
    print(f"  Budget:          ${budget['total']:>10,.0f} total | "
          f"${budget['spent']:>10,.0f} spent | "
          f"${budget_remaining:>10,.0f} remaining")
    print(f"  Collection:      {len(collection)} guitars")
    print(f"  Active listings: {len(listings)}")
    print(f"  Scoring model:   {dims} (weights: {json.dumps(w)})")

    # ML status line
    if budget.get("ml_enabled") and _ml_predict and _ml_predict.is_ml_available():
        blend = budget.get("ml_blend", 0.3)
        n_models = _ml_predict.active_model_count()
        print(f"  ML scoring:      {dims} hybrid (blend: {blend}, models: {n_models}/4)")
    elif budget.get("ml_enabled"):
        print(f"  ML scoring:      enabled but no models trained yet")
    else:
        print(f"  ML scoring:      disabled")

    print(f"  Scoring …\n")

    scored = []
    for entry in listings:
        total, breakdown = score_listing(entry, collection, budget)
        scored.append((entry, total, breakdown))

    scored.sort(key=lambda x: x[1], reverse=True)
    top = scored[:budget["top_n"]]

    # ── terminal summary ──
    print(f"  {'#':<4} {'Brand':<22} {'Model':<24} {'Year':>6} "
          f"{'Price':>10} {'Score':>6} {'Cond':>5} {'Icon':>5}  Budget")
    print(f"  {'─'*101}")
    for rank, (entry, total, bd) in enumerate(top, 1):
        price_str  = f"${entry['price']:>9,.0f}" if entry["price"] else "          N/A"
        budget_tag = "OVER" if (entry["price"] or 0) > budget_remaining else "ok"
        print(f"  {rank:<4} {entry['brand']:<22} {entry['model']:<24} "
              f"{entry['year']:>6} {price_str} {total:>6.1f} {bd['condition']:>5.0f} "
              f"{bd['iconic']:>5.0f}  {budget_tag}")
    print()

    write_recommendations(top, budget)
    print(f"  'Recommendations' sheet written to listings.xlsx")


if __name__ == "__main__":
    run()
