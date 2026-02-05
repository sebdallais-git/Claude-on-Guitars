#!/usr/bin/env python3
"""
watchdog — two jobs in one loop:

  1. keeps searcher.py alive (restarts it on exit)
  2. watches just-arrived-retrofret.xlsx for new rows and emails each one
         to sebdallais@gmail.com

Gmail requires an **App Password** (not your real password).
Create one at https://myaccount.google.com/apppasswords
then either:
    export GMAIL_APP_PASSWORD="your-16-char-code"
    python3 watchdog.py
or just run it — you'll be prompted at startup.
"""

import getpass
import json
import os
import re
import smtplib
import subprocess
import sys
import time
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, date
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import requests

try:
    from openpyxl import load_workbook
except ImportError:
    print("Missing dep — run:  pip install openpyxl")
    sys.exit(1)

import messenger           # Telegram (optional — needs env vars; see setup)

# ── config ────────────────────────────────────────────────────────
_SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.dirname(_SCRIPT_DIR)
_DATA         = os.path.join(_PROJECT_ROOT, "data")
_OUTPUTS      = os.path.join(_PROJECT_ROOT, "outputs")

SEARCHER        = os.path.join(_SCRIPT_DIR, "searcher.py")
XLSX_PATH       = os.path.join(_OUTPUTS, "listings.xlsx")
NOTIFIED_FILE      = os.path.join(_DATA, ".notified.json")        # new-entry emails sent
NOTIFIED_HOLD_FILE = os.path.join(_DATA, ".notified-hold.json")   # on-hold emails sent
NOTIFIED_SOLD_FILE = os.path.join(_DATA, ".notified-sold.json")   # sold  emails sent

CHECK_INTERVAL  = 15      # seconds between polls
RECIPIENT       = "sebdallais@gmail.com"
SMTP_HOST       = "smtp.gmail.com"
SMTP_PORT       = 587

RETROFRET_BASE  = "https://www.retrofret.com"
HTTP_HEADERS    = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    )
}


# ── helpers ───────────────────────────────────────────────────────
def log(msg):
    print(f"[{datetime.now():%Y-%m-%d %H:%M:%S}]  watchdog — {msg}", flush=True)


def fmt_price(raw):
    """Turn a float/int from openpyxl into '$1,234.00'."""
    if raw is None:
        return "N/A"
    try:
        return f"${float(raw):,.2f}"
    except (ValueError, TypeError):
        return str(raw)


# ── notified-ID persistence ───────────────────────────────────────
def load_notified():
    if os.path.exists(NOTIFIED_FILE):
        with open(NOTIFIED_FILE) as f:
            return set(json.load(f))
    return set()


def save_notified(ids):
    with open(NOTIFIED_FILE, "w") as f:
        json.dump(sorted(ids), f)


def load_notified_hold():
    if os.path.exists(NOTIFIED_HOLD_FILE):
        with open(NOTIFIED_HOLD_FILE) as f:
            return set(json.load(f))
    return set()


def save_notified_hold(ids):
    with open(NOTIFIED_HOLD_FILE, "w") as f:
        json.dump(sorted(ids), f)


def load_notified_sold():
    if os.path.exists(NOTIFIED_SOLD_FILE):
        with open(NOTIFIED_SOLD_FILE) as f:
            return set(json.load(f))
    return set()


def save_notified_sold(ids):
    with open(NOTIFIED_SOLD_FILE, "w") as f:
        json.dump(sorted(ids), f)


# ── xlsx reader ───────────────────────────────────────────────────
def read_entries():
    """Return list of row-dicts from the xlsx.  Safe to call while searcher writes."""
    if not os.path.exists(XLSX_PATH):
        return []
    try:
        wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(min_row=2, max_col=12, values_only=True):
            url = str(row[9]) if row[9] else ""
            m   = re.search(r"ProductID=(\d+)", url)
            if not m:
                continue
            date_raw = row[0]
            if isinstance(date_raw, (datetime, date)):
                date_arrived = date_raw.date() if isinstance(date_raw, datetime) else date_raw
            elif date_raw:
                try:
                    date_arrived = date.fromisoformat(str(date_raw)[:10])
                except (ValueError, TypeError):
                    date_arrived = None
            else:
                date_arrived = None

            rows.append({
                "id":           m.group(1),
                "date_arrived": date_arrived,
                "brand":        str(row[1] or ""),
                "model":        str(row[2] or ""),
                "type":         str(row[3] or ""),
                "year":         str(row[4] or ""),
                "price":        fmt_price(row[5]),
                "reverb_low":   fmt_price(row[6]),
                "reverb_hi":    fmt_price(row[7]),
                "condition":    str(row[8] or ""),
                "url":          url,
                "on_hold":      row[10] is not None,
                "sold":         row[11] is not None,
            })
        wb.close()
        return rows
    except Exception as e:
        log(f"xlsx read error (will retry): {e}")
        return []


# ── image ─────────────────────────────────────────────────────────
def fetch_listing_image(pid):
    """Download the main product photo.  Returns (bytes, subtype) or (None, None)."""
    url = f"{RETROFRET_BASE}/images/{pid}_Guitar/{pid}_01.jpg"
    try:
        resp = requests.get(url, headers=HTTP_HEADERS, timeout=10)
        resp.raise_for_status()
        subtype = resp.headers.get("Content-Type", "image/jpeg").split("/")[-1]
        return resp.content, subtype
    except Exception as e:
        log(f"image fetch failed ({pid}): {e}")
        return None, None


# ── email ─────────────────────────────────────────────────────────
def _build_msg(entry, on_hold=False, img_data=None, img_subtype="jpeg"):
    tag     = " ON HOLD" if on_hold else ""
    subject = (
        f"[RetroFret{tag}] {entry['brand']} {entry['model']} "
        f"{entry['year']} — {entry['price']} — {entry['condition']}"
    )
    details = (
        f"Brand:     {entry['brand']}\n"
        f"Model:     {entry['model']}\n"
        f"Year:      {entry['year']}\n"
        f"Type:      {entry['type']}\n"
        f"Price:     {entry['price']}\n" +
        (f"Reverb:    {entry['reverb_low']} – {entry['reverb_hi']}\n"
         if entry.get("reverb_low") and entry["reverb_low"] != "N/A" else "") +
        f"Condition: {entry['condition']}\n"
        f"\n"
        f"{entry['url']}\n"
    )

    if img_data:
        hold_html = "<p><strong>— ON HOLD —</strong></p>\n" if on_hold else ""
        html = (
            "<html><body>\n"
            f"{hold_html}"
            f'<img src="cid:listing_photo" style="max-width:600px"><br>\n'
            f"<pre>{details}</pre>\n"
            "</body></html>\n"
        )
        msg = MIMEMultipart("related")
        msg.attach(MIMEText(html, "html"))
        img_part = MIMEImage(img_data, _subtype=img_subtype)
        img_part.add_header("Content-Disposition", "inline")
        img_part.add_header("Content-ID", "<listing_photo>")
        msg.attach(img_part)
    else:
        body = (f"— on hold —\n\n" if on_hold else "") + details
        msg  = MIMEText(body, "plain")

    msg["From"]    = RECIPIENT
    msg["To"]      = RECIPIENT
    msg["Subject"] = subject
    return msg


def _build_sold_msg(entry):
    subject = (
        f"[RetroFret SOLD] {entry['brand']} {entry['model']} "
        f"{entry['year']} — {entry['price']}"
    )
    details = (
        f"Brand:     {entry['brand']}\n"
        f"Model:     {entry['model']}\n"
        f"Year:      {entry['year']}\n"
        f"Type:      {entry['type']}\n"
        f"Price:     {entry['price']}\n" +
        (f"Reverb:    {entry['reverb_low']} – {entry['reverb_hi']}\n"
         if entry.get("reverb_low") and entry["reverb_low"] != "N/A" else "") +
        f"Condition: {entry['condition']}\n"
        f"\n{entry['url']}\n"
    )
    msg            = MIMEText(f"— SOLD —\n\n{details}", "plain")
    msg["From"]    = RECIPIENT
    msg["To"]      = RECIPIENT
    msg["Subject"] = subject
    return msg


# ── searcher launcher ─────────────────────────────────────────────
def launch_searcher():
    log("launching searcher.py …")
    return subprocess.Popen(
        [sys.executable, "-u", SEARCHER],
        cwd=_SCRIPT_DIR,
        stdout=sys.stdout,
        stderr=sys.stderr,
    )


# ── main loop ─────────────────────────────────────────────────────
def main():
    # grab app password once, up front
    app_pw = os.environ.get("GMAIL_APP_PASSWORD")
    if not app_pw:
        log("no GMAIL_APP_PASSWORD env var — prompting (input hidden)")
        app_pw = getpass.getpass("  Gmail App Password: ").strip()
    if not app_pw:
        log("no password provided — email notifications disabled")

    if messenger.enabled():
        log("Telegram enabled")
    else:
        log("Telegram disabled — set TELEGRAM_BOT_TOKEN + TELEGRAM_CHAT_ID to enable")

    log("started")
    proc = launch_searcher()

    # seed on first run so we don't blast everything already in the xlsx
    first_run      = not os.path.exists(NOTIFIED_FILE)
    first_run_hold = not os.path.exists(NOTIFIED_HOLD_FILE)
    first_run_sold = not os.path.exists(NOTIFIED_SOLD_FILE)
    notified       = load_notified()
    notified_hold  = load_notified_hold()
    notified_sold  = load_notified_sold()
    if first_run or first_run_hold or first_run_sold:
        entries = read_entries()
        if first_run:
            notified = {e["id"] for e in entries}
            save_notified(notified)
        if first_run_hold:
            notified_hold = {e["id"] for e in entries if e.get("on_hold")}
            save_notified_hold(notified_hold)
        if first_run_sold:
            notified_sold = {e["id"] for e in entries if e.get("sold")}
            save_notified_sold(notified_sold)
        log(f"seeded — {len(notified)} entries, "
            f"{len(notified_hold)} on hold, {len(notified_sold)} sold")

    while True:
        time.sleep(CHECK_INTERVAL)

        # --- 1. keep searcher alive ---
        if proc.poll() is not None:
            log(f"searcher.py exited (code {proc.returncode}) — restarting …")
            proc = launch_searcher()

        # --- 2. detect changes in xlsx ---
        today        = date.today()
        new_pending  = []
        hold_pending = []
        sold_pending = []
        for entry in read_entries():
            if entry["id"] not in notified:
                if entry["date_arrived"] == today:
                    new_pending.append(entry)
                else:
                    notified.add(entry["id"])   # too old — mark seen silently
            if entry.get("on_hold") and entry["id"] not in notified_hold:
                hold_pending.append(entry)
            if entry.get("sold") and entry["id"] not in notified_sold:
                sold_pending.append(entry)
        save_notified(notified)                 # persist any silently-added IDs

        # --- 3. notify new + on-hold ---
        pending = [(e, False) for e in new_pending] + [(e, True) for e in hold_pending]
        if pending:
            # fetch listing photos in parallel
            with ThreadPoolExecutor(max_workers=5) as pool:
                img_futures = [pool.submit(fetch_listing_image, e["id"]) for e, _ in pending]
                images      = [f.result() for f in img_futures]

            # email batch
            if app_pw:
                try:
                    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as srv:
                        srv.starttls()
                        srv.login(RECIPIENT, app_pw)
                        for (entry, on_hold), (img_data, img_sub) in zip(pending, images):
                            try:
                                srv.send_message(_build_msg(entry, on_hold, img_data, img_sub or "jpeg"))
                                action = "on hold" if on_hold else "emailed"
                                log(f"{action} — {entry['brand']} {entry['model']} {entry['year']}")
                            except Exception as e:
                                log(f"send failed (will retry next cycle): {e}")
                                break  # connection dead
                except smtplib.SMTPAuthenticationError:
                    log("Gmail rejected the App Password — email disabled for this session")
                    log("Generate a new one at https://myaccount.google.com/apppasswords")
                    app_pw = None
                except Exception as e:
                    log(f"SMTP session failed (will retry): {e}")

            # Telegram batch
            if messenger.enabled():
                for entry, on_hold in pending:
                    fn = messenger.notify_on_hold if on_hold else messenger.notify_new
                    ok = fn(entry)
                    tag = "on hold" if on_hold else "new"
                    log(f"telegram {tag} — {entry['brand']} {entry['model']} "
                        f"{entry['year']} {'ok' if ok else 'FAILED'}")

            # mark as notified (after both channels attempted)
            for entry, on_hold in pending:
                if on_hold:
                    notified_hold.add(entry["id"])
                else:
                    notified.add(entry["id"])
            save_notified(notified)
            save_notified_hold(notified_hold)

        # --- 4. notify sold ---
        if sold_pending:
            if app_pw:
                try:
                    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as srv:
                        srv.starttls()
                        srv.login(RECIPIENT, app_pw)
                        for entry in sold_pending:
                            try:
                                srv.send_message(_build_sold_msg(entry))
                                log(f"emailed sold — {entry['brand']} {entry['model']} {entry['year']}")
                            except Exception as e:
                                log(f"sold email failed: {e}")
                                break
                except smtplib.SMTPAuthenticationError:
                    log("Gmail rejected the App Password — email disabled")
                    app_pw = None
                except Exception as e:
                    log(f"SMTP session failed (will retry): {e}")

            if messenger.enabled():
                for entry in sold_pending:
                    ok = messenger.notify_sold(entry)
                    log(f"telegram sold — {entry['brand']} {entry['model']} "
                        f"{entry['year']} {'ok' if ok else 'FAILED'}")

            for entry in sold_pending:
                notified_sold.add(entry["id"])
            save_notified_sold(notified_sold)


if __name__ == "__main__":
    main()
