#!/usr/bin/env python3
"""
dashboard — web server for viewing listings and recommendations on any device.

Serves the dashboard on your local network so you can view it from iPad/phone.
Access via: http://YOUR_IP:8080

Usage:
    python3 scripts/dashboard.py
"""

import os
import json
from http.server import HTTPServer, SimpleHTTPRequestHandler
from openpyxl import load_workbook
from urllib.parse import urlparse


_SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.dirname(_SCRIPT_DIR)
_DATA         = os.path.join(_PROJECT_ROOT, "data")
_OUTPUTS      = os.path.join(_PROJECT_ROOT, "outputs")

XLSX_PATH         = os.path.join(_OUTPUTS, "listings.xlsx")
COLLECTION_FILE   = os.path.join(_DATA, "collection.json")
BUDGET_FILE       = os.path.join(_DATA, "budget.json")
BRAND_TIERS_FILE    = os.path.join(_DATA, "knowledge", "brand_tiers.json")
ICONIC_MODELS_FILE  = os.path.join(_DATA, "knowledge", "iconic_models.json")
TOP_GUITARISTS_FILE = os.path.join(_DATA, "knowledge", "top_guitarists.json")
PRICE_HISTORY_FILE  = os.path.join(_DATA, "price_history.json")
_ML_DIR             = os.path.join(_DATA, "ml")
ML_TRAINING_FILE    = os.path.join(_ML_DIR, "training_data.json")
ML_PERFORMANCE_FILE = os.path.join(_ML_DIR, "performance.json")


def xlsx_to_json(sheet_name="Sheet"):
    """Convert an Excel sheet to JSON."""
    if not os.path.exists(XLSX_PATH):
        return []

    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return []

    headers = rows[0]
    data = []
    for row in rows[1:]:
        entry = {}
        for i, header in enumerate(headers):
            entry[header] = row[i] if i < len(row) else None
        data.append(entry)

    wb.close()
    return data


class DashboardHandler(SimpleHTTPRequestHandler):
    """Custom handler that serves API endpoints and static files."""

    def do_GET(self):
        parsed_path = urlparse(self.path)
        path = parsed_path.path

        # API endpoints
        if path == "/api/listings":
            self.send_json_response(xlsx_to_json("Guitars"))
        elif path == "/api/recommendations":
            self.send_json_response(xlsx_to_json("Recommendations"))
        elif path == "/api/collection":
            self.send_json_response(self.load_json_file(COLLECTION_FILE))
        elif path == "/api/budget":
            self.send_json_response(self.load_json_file(BUDGET_FILE))
        elif path == "/api/stats":
            self.send_json_response(self.get_stats())
        elif path == "/api/agent-status":
            self.send_json_response(self.get_agent_status())
        elif path == "/api/scorer-info":
            self.send_json_response(self.get_scorer_info())
        elif path == "/api/iconic-models":
            self.send_json_response(self.get_iconic_models())
        elif path == "/api/top-guitarists":
            self.send_json_response(self.get_top_guitarists())
        elif path == "/api/ml-status":
            self.send_json_response(self.get_ml_status())
        elif path == "/api/ml-performance":
            self.send_json_response(self.get_ml_performance())
        elif path == "/" or path == "/index.html":
            self.send_dashboard_html()
        else:
            # Serve static files from project root
            self.path = path
            super().do_GET()

    def do_POST(self):
        parsed_path = urlparse(self.path)
        path = parsed_path.path

        if path == "/api/budget":
            self.update_budget()
        else:
            self.send_error(404, "Not Found")

    def do_OPTIONS(self):
        """Handle CORS preflight requests."""
        self.send_response(204)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()

    def send_json_response(self, data):
        """Send JSON response."""
        self.send_response(200)
        self.send_header("Content-Type", "application/json")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        self.wfile.write(json.dumps(data, default=str).encode())

    def load_json_file(self, filepath):
        """Load JSON file or return empty array/object."""
        if not os.path.exists(filepath):
            return []
        try:
            with open(filepath) as f:
                return json.load(f)
        except:
            return []

    def get_stats(self):
        """Calculate dashboard statistics."""
        listings = xlsx_to_json("Guitars")
        recommendations = xlsx_to_json("Recommendations")
        collection = self.load_json_file(COLLECTION_FILE)
        budget = self.load_json_file(BUDGET_FILE)

        active_listings = sum(1 for l in listings if not l.get("Sold") and not l.get("On Hold"))

        return {
            "total_listings": len(listings),
            "active_listings": active_listings,
            "recommendations": len(recommendations),
            "collection_size": len(collection),
            "budget_total": budget.get("total", 0) if budget else 0,
            "budget_spent": budget.get("spent", 0) if budget else 0,
            "budget_remaining": (budget.get("total", 0) - budget.get("spent", 0)) if budget else 0,
        }

    def send_dashboard_html(self):
        """Serve the main dashboard HTML."""
        self.send_response(200)
        self.send_header("Content-Type", "text/html")
        self.end_headers()

        html = open(os.path.join(_PROJECT_ROOT, "dashboard.html"), "rb").read()
        self.wfile.write(html)

    def get_scorer_info(self):
        """Return scorer metadata: weights, knowledge base stats, features, and sources."""
        budget = self.load_json_file(BUDGET_FILE)
        if isinstance(budget, list):
            budget = {}
        weights = budget.get("weights", {
            "value": 0.30, "appreciate": 0.25, "fit": 0.25, "condition": 0.20
        })

        # Count knowledge base entries
        brand_tiers = self.load_json_file(BRAND_TIERS_FILE)
        if isinstance(brand_tiers, list):
            brand_tiers = {}
        premium_count = len(brand_tiers.get("premium", []))
        major_count = len(brand_tiers.get("major", []))

        iconic = self.load_json_file(ICONIC_MODELS_FILE)
        if isinstance(iconic, list):
            iconic = {}
        iconic_count = len(iconic.get("models", []))

        top_guitarists = self.load_json_file(TOP_GUITARISTS_FILE)
        if isinstance(top_guitarists, list):
            top_guitarists = {}
        guitarists_count = len(top_guitarists.get("guitarists", []))

        price_history = self.load_json_file(PRICE_HISTORY_FILE)
        if isinstance(price_history, list):
            price_history = {}
        snapshots = price_history.get("snapshots", {})
        snapshot_models = len(snapshots)
        total_snapshots = sum(len(v) for v in snapshots.values())

        # Count learned rates (models with 2+ snapshots spanning 30+ days)
        learned_rates = 0
        for entries in snapshots.values():
            if len(entries) >= 2:
                try:
                    from datetime import datetime
                    first = datetime.strptime(entries[0]["date"], "%Y-%m-%d")
                    last = datetime.strptime(entries[-1]["date"], "%Y-%m-%d")
                    if (last - first).days >= 30:
                        learned_rates += 1
                except (KeyError, ValueError):
                    pass

        return {
            "weights": weights,
            "dimensions": len(weights),
            "knowledge_base": {
                "iconic_models": iconic_count,
                "top_guitarists": guitarists_count,
                "brand_tiers": {"premium": premium_count, "major": major_count},
                "learned_rates": learned_rates,
                "price_snapshots": total_snapshots,
                "snapshot_models": snapshot_models
            },
            "top_features": [
                {"name": "Value opportunity", "weight": weights.get("value", 0.25),
                 "description": "Price vs Reverb market range"},
                {"name": "Appreciation potential", "weight": weights.get("appreciate", 0.20),
                 "description": "Era + brand tier + learned rates"},
                {"name": "Collection fit", "weight": weights.get("fit", 0.20),
                 "description": "Diversification + iconic model boost"},
                {"name": "Condition", "weight": weights.get("condition", 0.20),
                 "description": "13-level condition scale (Mint to Poor)"},
                {"name": "Iconic status", "weight": weights.get("iconic", 0.15),
                 "description": "Top-100 guitarist associations (rank-weighted)"}
            ],
            "boosts": [
                {"name": "Golden era", "points": "+20 appreciation", "source": "iconic_models.json"},
                {"name": "Iconic model", "points": "+0-20 fit", "source": "iconic_models.json"},
                {"name": "New brand", "points": "+20 fit", "source": "collection.json"},
                {"name": "Rare type", "points": "+15 fit", "source": "collection.json"}
            ],
            "information_sources": [
                {"name": "Reverb Price Guide", "type": "external", "usage": "Value scoring, price snapshots"},
                {"name": "brand_tiers.json", "type": "knowledge", "usage": "3-tier brand classification"},
                {"name": "iconic_models.json", "type": "knowledge", "usage": "Golden era + popularity boosts"},
                {"name": "top_guitarists.json", "type": "knowledge", "usage": "Top 100 guitarists and their guitars"},
                {"name": "price_history.json", "type": "learned", "usage": "Learned appreciation rates"},
                {"name": "collection.json", "type": "user", "usage": "Fit scoring (diversification)"},
                {"name": "budget.json", "type": "config", "usage": "Weights, budget limits, top_n"}
            ],
            "condition_scale": [
                {"label": "Mint", "score": 100},
                {"label": "Near Mint", "score": 95},
                {"label": "Excellent", "score": 85},
                {"label": "Very Good", "score": 60},
                {"label": "Good", "score": 30},
                {"label": "Poor", "score": 0}
            ]
        }

    def get_ml_status(self):
        """Return ML system status: model availability, config, training info."""
        budget = self.load_json_file(BUDGET_FILE)
        if isinstance(budget, list):
            budget = {}

        training = self.load_json_file(ML_TRAINING_FILE)
        if isinstance(training, list):
            training = {}
        sold_count = len(training.get("sold_listings", []))

        # Check each model's metadata
        model_names = [
            "weight_optimizer", "price_predictor",
            "appreciation_predictor", "buy_classifier",
        ]
        models = {}
        last_trained = None
        for name in model_names:
            meta_path = os.path.join(_ML_DIR, "models", f"{name}_meta.json")
            meta = self.load_json_file(meta_path)
            if isinstance(meta, list):
                meta = {}
            available = os.path.exists(
                os.path.join(_ML_DIR, "models", f"{name}.joblib")
            )
            models[name] = {
                "available": available,
                "trained_at": meta.get("trained_at"),
                "samples": meta.get("samples", 0),
                "metrics": meta.get("metrics", {}),
            }
            if meta.get("trained_at"):
                if last_trained is None or meta["trained_at"] > last_trained:
                    last_trained = meta["trained_at"]

        # Days of data from performance logs
        perf = self.load_json_file(ML_PERFORMANCE_FILE)
        if isinstance(perf, list):
            perf = {}
        days = len(perf.get("daily_logs", []))

        return {
            "ml_enabled": budget.get("ml_enabled", False),
            "ml_blend": budget.get("ml_blend", 0.3),
            "models_active": sum(1 for m in models.values() if m["available"]),
            "models": models,
            "training_samples": sold_count,
            "days_of_data": days,
            "last_trained": last_trained,
        }

    def get_ml_performance(self):
        """Return last 30 daily performance logs."""
        perf = self.load_json_file(ML_PERFORMANCE_FILE)
        if isinstance(perf, list):
            perf = {}
        logs = perf.get("daily_logs", [])
        return {"daily_logs": logs[-30:]}

    def update_budget(self):
        """Update budget total and spent from POST JSON body."""
        try:
            length = int(self.headers.get("Content-Length", 0))
            body = json.loads(self.rfile.read(length))
        except (ValueError, json.JSONDecodeError):
            self.send_error(400, "Invalid JSON")
            return

        total = body.get("total")
        spent = body.get("spent")

        # Validate: both must be numbers, non-negative, spent <= total
        if not isinstance(total, (int, float)) or not isinstance(spent, (int, float)):
            self.send_error(400, "total and spent must be numbers")
            return
        if total < 0 or spent < 0:
            self.send_error(400, "total and spent must be non-negative")
            return
        if spent > total:
            self.send_error(400, "spent cannot exceed total")
            return

        # Load existing budget.json, update only total/spent
        budget = self.load_json_file(BUDGET_FILE)
        if isinstance(budget, list):
            budget = {}
        budget["total"] = total
        budget["spent"] = spent

        with open(BUDGET_FILE, "w") as f:
            json.dump(budget, f, indent=4)

        self.send_json_response(budget)

    def get_iconic_models(self):
        """Return iconic models with value ranges from price history."""
        iconic = self.load_json_file(ICONIC_MODELS_FILE)
        if isinstance(iconic, list):
            iconic = {}
        models = iconic.get("models", [])

        # Load price history for value ranges
        price_history = self.load_json_file(PRICE_HISTORY_FILE)
        if isinstance(price_history, list):
            price_history = {}
        snapshots = price_history.get("snapshots", {})

        result = []
        for m in models:
            brand = m.get("brand", "")
            model = m.get("model", "")
            golden_era = m.get("golden_era", [])
            artists = m.get("artists", [])
            boost = m.get("boost", 0)

            # Look up value range: key is "brand.lower()|model.lower()"
            key = brand.lower() + "|" + model.lower()
            value_lo = None
            value_hi = None
            if key in snapshots and snapshots[key]:
                latest = snapshots[key][-1]
                value_lo = latest.get("reverb_lo")
                value_hi = latest.get("reverb_hi")

            result.append({
                "brand": brand,
                "model": model,
                "golden_era": golden_era,
                "boost": boost,
                "artists": artists[:3],
                "value_lo": value_lo,
                "value_hi": value_hi,
            })

        # Sort by boost descending
        result.sort(key=lambda x: x["boost"], reverse=True)
        return result

    def get_top_guitarists(self):
        """Return top guitarists list as-is (already sorted by rank)."""
        data = self.load_json_file(TOP_GUITARISTS_FILE)
        if isinstance(data, list):
            data = {}
        return data.get("guitarists", [])

    def get_agent_status(self):
        """Get status of running agents."""
        import subprocess
        import time

        status = {}

        # Check watchdog
        try:
            result = subprocess.run(
                ['ps', 'aux'],
                capture_output=True,
                text=True,
                timeout=5
            )

            watchdog_running = False
            searcher_running = False
            watchdog_pid = None

            for line in result.stdout.split('\n'):
                if 'watchdog.py' in line and 'grep' not in line:
                    watchdog_running = True
                    parts = line.split()
                    if len(parts) > 1:
                        watchdog_pid = parts[1]
                if 'searcher.py' in line and 'grep' not in line:
                    searcher_running = True

            if watchdog_running and watchdog_pid:
                status['watchdog'] = {
                    'running': True,
                    'uptime': 3600,
                    'last_check': time.strftime('%Y-%m-%d %H:%M:%S')
                }


            status['scraper'] = {
                'running': searcher_running
            }

        except Exception as e:
            status = {
                'error': str(e),
                'watchdog': {'running': False},
                'scraper': {'running': False}
            }

        return status


def main():
    """Start the dashboard server."""
    PORT = 8080

    # Change to project root so relative paths work
    os.chdir(_PROJECT_ROOT)

    server = HTTPServer(("0.0.0.0", PORT), DashboardHandler)

    # Get local IP
    import socket
    hostname = socket.gethostname()
    local_ip = socket.gethostbyname(hostname)

    print(f"""
╔══════════════════════════════════════════════════════════╗
║  Vintage Guitar Collector Dashboard                     ║
╚══════════════════════════════════════════════════════════╝

  Local:   http://localhost:{PORT}
  Network: http://{local_ip}:{PORT}

  📱 Access from your iPad using the Network URL
  🛑 Press Ctrl+C to stop the server

""")

    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n\n  Server stopped.")
        server.shutdown()


if __name__ == "__main__":
    main()
